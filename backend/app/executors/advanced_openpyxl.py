# -*- coding: utf-8 -*-
"""Excel 自动化模块执行器（基于 openpyxl）

提供一整套对 .xlsx 表格的自动化操作：工作簿/工作表管理、单元格与区域读写、
行列增删、样式设置、合并单元格、列宽行高、公式、查找替换、清空等。

设计原则（与项目其它模块一致）：
- 文件路径式、无状态：每个模块打开文件 → 操作 → 保存，互不依赖会话句柄
- openpyxl 懒加载（函数内 import），不拖慢后端启动
- 所有配置值都过 context.resolve_value，支持 {变量名}
"""
import os
import json
from typing import Any

from .base import ModuleExecutor, ExecutionContext, ModuleResult, register_executor
from .type_utils import to_int, to_bool


def _openpyxl():
    import openpyxl
    return openpyxl


def _resolve_path(context: ExecutionContext, raw: str) -> str:
    """解析文件路径：绝对路径原样；相对路径相对当前工作目录。"""
    p = (raw or "").strip().strip('"').strip("'")
    return p


def _load_wb(path: str, read_only: bool = False):
    """加载工作簿（不存在则报错）。"""
    openpyxl = _openpyxl()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel 文件不存在: {path}")
    return openpyxl.load_workbook(path, read_only=read_only, data_only=False)


def _get_ws(wb, sheet_name: str):
    """按名称取工作表；为空取活动表；不存在报错。"""
    if not sheet_name:
        return wb.active
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}（现有: {', '.join(wb.sheetnames)}）")
    return wb[sheet_name]


def _parse_2d(raw: Any) -> list:
    """把 JSON 字符串 / list 解析成二维数组。一维则包成一行。"""
    if isinstance(raw, (list, tuple)):
        data = list(raw)
    else:
        s = str(raw or "").strip()
        if not s:
            return []
        data = json.loads(s)
    if not isinstance(data, list):
        return [[data]]
    # 一维 → 单行
    if data and not isinstance(data[0], (list, tuple)):
        return [data]
    return [list(r) for r in data]


def _parse_1d(raw: Any) -> list:
    """把 JSON 字符串 / list 解析成一维数组。"""
    if isinstance(raw, (list, tuple)):
        return list(raw)
    s = str(raw or "").strip()
    if not s:
        return []
    try:
        v = json.loads(s)
    except Exception:
        # 逗号分隔兜底
        return [x.strip() for x in s.split(",")]
    if isinstance(v, list):
        return v
    return [v]


@register_executor
class ExcelCreateExecutor(ModuleExecutor):
    """创建 Excel 工作簿"""

    @property
    def module_type(self) -> str:
        return "excel_create"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_names_raw = context.resolve_value(config.get("sheetNames", "Sheet1")) or "Sheet1"
        overwrite = to_bool(config.get("overwrite", False), context)
        if not path:
            return ModuleResult(success=False, error="文件路径不能为空")
        if os.path.exists(path) and not overwrite:
            return ModuleResult(success=False, error=f"文件已存在（如需覆盖请开启 overwrite）: {path}")
        try:
            openpyxl = _openpyxl()
            names = [n.strip() for n in str(sheet_names_raw).split(",") if n.strip()] or ["Sheet1"]
            wb = openpyxl.Workbook()
            wb.active.title = names[0]
            for n in names[1:]:
                wb.create_sheet(title=n)
            os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
            wb.save(path)
            return ModuleResult(success=True, message=f"已创建 Excel: {path}（工作表: {', '.join(names)}）",
                                data={"path": path, "sheets": names})
        except Exception as e:
            return ModuleResult(success=False, error=f"创建 Excel 失败: {e}")


@register_executor
class ExcelAddSheetExecutor(ModuleExecutor):
    """添加工作表"""

    @property
    def module_type(self) -> str:
        return "excel_add_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", "")) or "新工作表"
        try:
            wb = _load_wb(path)
            if sheet_name in wb.sheetnames:
                return ModuleResult(success=False, error=f"工作表已存在: {sheet_name}")
            wb.create_sheet(title=sheet_name)
            wb.save(path)
            return ModuleResult(success=True, message=f"已添加工作表: {sheet_name}", data={"sheets": wb.sheetnames})
        except Exception as e:
            return ModuleResult(success=False, error=f"添加工作表失败: {e}")


@register_executor
class ExcelDeleteSheetExecutor(ModuleExecutor):
    """删除工作表"""

    @property
    def module_type(self) -> str:
        return "excel_delete_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        try:
            wb = _load_wb(path)
            if sheet_name not in wb.sheetnames:
                return ModuleResult(success=False, error=f"工作表不存在: {sheet_name}")
            if len(wb.sheetnames) <= 1:
                return ModuleResult(success=False, error="工作簿至少要保留一个工作表")
            del wb[sheet_name]
            wb.save(path)
            return ModuleResult(success=True, message=f"已删除工作表: {sheet_name}", data={"sheets": wb.sheetnames})
        except Exception as e:
            return ModuleResult(success=False, error=f"删除工作表失败: {e}")


@register_executor
class ExcelRenameSheetExecutor(ModuleExecutor):
    """重命名工作表"""

    @property
    def module_type(self) -> str:
        return "excel_rename_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        old_name = context.resolve_value(config.get("oldName", ""))
        new_name = context.resolve_value(config.get("newName", ""))
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, old_name)
            ws.title = new_name
            wb.save(path)
            return ModuleResult(success=True, message=f"已重命名: {old_name} → {new_name}", data={"sheets": wb.sheetnames})
        except Exception as e:
            return ModuleResult(success=False, error=f"重命名工作表失败: {e}")


@register_executor
class ExcelListSheetsExecutor(ModuleExecutor):
    """列出所有工作表名"""

    @property
    def module_type(self) -> str:
        return "excel_list_sheets"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            wb = _load_wb(path, read_only=True)
            names = list(wb.sheetnames)
            wb.close()
            if var:
                context.set_variable(var, names)
            return ModuleResult(success=True, message=f"共 {len(names)} 个工作表: {', '.join(names)}", data={"sheets": names})
        except Exception as e:
            return ModuleResult(success=False, error=f"列出工作表失败: {e}")


@register_executor
class ExcelWriteCellExecutor(ModuleExecutor):
    """写入单元格"""

    @property
    def module_type(self) -> str:
        return "excel_write_cell"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "")) or "").strip().upper()
        value = context.resolve_value(config.get("value", ""))
        if not cell:
            return ModuleResult(success=False, error="单元格地址不能为空（如 A1）")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws[cell] = value
            wb.save(path)
            return ModuleResult(success=True, message=f"已写入 {cell} = {value}")
        except Exception as e:
            return ModuleResult(success=False, error=f"写入单元格失败: {e}")


@register_executor
class ExcelReadCellExecutor(ModuleExecutor):
    """读取单元格"""

    @property
    def module_type(self) -> str:
        return "excel_read_cell"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "")) or "").strip().upper()
        var = config.get("resultVariable", "") or config.get("variableName", "")
        if not cell:
            return ModuleResult(success=False, error="单元格地址不能为空（如 A1）")
        try:
            wb = _load_wb(path, read_only=True)
            ws = _get_ws(wb, sheet_name)
            value = ws[cell].value
            wb.close()
            if var:
                context.set_variable(var, value)
            return ModuleResult(success=True, message=f"{cell} = {value}", data={"value": value})
        except Exception as e:
            return ModuleResult(success=False, error=f"读取单元格失败: {e}")


@register_executor
class ExcelWriteRangeExecutor(ModuleExecutor):
    """写入区域（二维数组，从起始单元格开始铺开）"""

    @property
    def module_type(self) -> str:
        return "excel_write_range"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        start_cell = (context.resolve_value(config.get("startCell", "A1")) or "A1").strip().upper()
        data_raw = context.resolve_value(config.get("data", ""))
        try:
            data = _parse_2d(data_raw)
            if not data:
                return ModuleResult(success=False, error="写入数据为空（需要二维数组 JSON，如 [[1,2],[3,4]]）")
            openpyxl = _openpyxl()
            from openpyxl.utils.cell import coordinate_to_tuple
            r0, c0 = coordinate_to_tuple(start_cell)  # (row, col)
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            n = 0
            for i, row in enumerate(data):
                for j, val in enumerate(row):
                    ws.cell(row=r0 + i, column=c0 + j, value=val)
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已写入 {len(data)} 行 × 共 {n} 格，起点 {start_cell}")
        except Exception as e:
            return ModuleResult(success=False, error=f"写入区域失败: {e}")


@register_executor
class ExcelReadRangeExecutor(ModuleExecutor):
    """读取区域（返回二维数组）"""

    @property
    def module_type(self) -> str:
        return "excel_read_range"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        var = config.get("resultVariable", "") or config.get("variableName", "")
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:C10）")
        try:
            wb = _load_wb(path, read_only=True)
            ws = _get_ws(wb, sheet_name)
            result = []
            for row in ws[cell_range]:
                # 单格时 ws[range] 返回单个 cell，统一成二维
                if not isinstance(row, tuple):
                    row = (row,)
                result.append([c.value for c in row])
            wb.close()
            if var:
                context.set_variable(var, result)
            return ModuleResult(success=True, message=f"已读取区域 {cell_range}（{len(result)} 行）", data={"data": result})
        except Exception as e:
            return ModuleResult(success=False, error=f"读取区域失败: {e}")


@register_executor
class ExcelAppendRowExecutor(ModuleExecutor):
    """追加一行到表格末尾"""

    @property
    def module_type(self) -> str:
        return "excel_append_row"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        row_raw = context.resolve_value(config.get("rowData", ""))
        try:
            row = _parse_1d(row_raw)
            if not row:
                return ModuleResult(success=False, error="行数据为空（需要数组 JSON，如 [\"张三\", 18, \"北京\"]）")
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.append(row)
            wb.save(path)
            return ModuleResult(success=True, message=f"已追加一行（{len(row)} 列），当前共 {ws.max_row} 行")
        except Exception as e:
            return ModuleResult(success=False, error=f"追加行失败: {e}")


@register_executor
class ExcelInsertRowsExecutor(ModuleExecutor):
    """插入空行（在指定行号之前插入 count 行）"""

    @property
    def module_type(self) -> str:
        return "excel_insert_rows"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        idx = to_int(config.get("rowIndex", 1), 1, context)
        count = to_int(config.get("count", 1), 1, context)
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.insert_rows(idx, amount=max(1, count))
            wb.save(path)
            return ModuleResult(success=True, message=f"已在第 {idx} 行前插入 {count} 行")
        except Exception as e:
            return ModuleResult(success=False, error=f"插入行失败: {e}")


@register_executor
class ExcelDeleteRowsExecutor(ModuleExecutor):
    """删除行（从指定行号开始删 count 行）"""

    @property
    def module_type(self) -> str:
        return "excel_delete_rows"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        idx = to_int(config.get("rowIndex", 1), 1, context)
        count = to_int(config.get("count", 1), 1, context)
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.delete_rows(idx, amount=max(1, count))
            wb.save(path)
            return ModuleResult(success=True, message=f"已从第 {idx} 行起删除 {count} 行")
        except Exception as e:
            return ModuleResult(success=False, error=f"删除行失败: {e}")


@register_executor
class ExcelInsertColsExecutor(ModuleExecutor):
    """插入空列（在指定列号之前插入 count 列）"""

    @property
    def module_type(self) -> str:
        return "excel_insert_cols"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        idx = to_int(config.get("colIndex", 1), 1, context)
        count = to_int(config.get("count", 1), 1, context)
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.insert_cols(idx, amount=max(1, count))
            wb.save(path)
            return ModuleResult(success=True, message=f"已在第 {idx} 列前插入 {count} 列")
        except Exception as e:
            return ModuleResult(success=False, error=f"插入列失败: {e}")


@register_executor
class ExcelDeleteColsExecutor(ModuleExecutor):
    """删除列（从指定列号开始删 count 列）"""

    @property
    def module_type(self) -> str:
        return "excel_delete_cols"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        idx = to_int(config.get("colIndex", 1), 1, context)
        count = to_int(config.get("count", 1), 1, context)
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.delete_cols(idx, amount=max(1, count))
            wb.save(path)
            return ModuleResult(success=True, message=f"已从第 {idx} 列起删除 {count} 列")
        except Exception as e:
            return ModuleResult(success=False, error=f"删除列失败: {e}")


@register_executor
class ExcelSetFormulaExecutor(ModuleExecutor):
    """设置单元格公式（如 =SUM(A1:A10)）"""

    @property
    def module_type(self) -> str:
        return "excel_set_formula"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "")) or "").strip().upper()
        formula = context.resolve_value(config.get("formula", ""))
        if not cell:
            return ModuleResult(success=False, error="单元格地址不能为空")
        try:
            f = str(formula or "")
            if f and not f.startswith("="):
                f = "=" + f
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws[cell] = f
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置公式 {cell} = {f}")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置公式失败: {e}")


@register_executor
class ExcelClearRangeExecutor(ModuleExecutor):
    """清空区域内容（保留格式）"""

    @property
    def module_type(self) -> str:
        return "excel_clear_range"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:C10）")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cells = ws[cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            n = 0
            for row in cells:
                if not isinstance(row, tuple):
                    row = (row,)
                for c in row:
                    c.value = None
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已清空区域 {cell_range}（{n} 格）")
        except Exception as e:
            return ModuleResult(success=False, error=f"清空区域失败: {e}")


@register_executor
class ExcelFindReplaceExecutor(ModuleExecutor):
    """查找替换（整表或指定工作表内文本替换）"""

    @property
    def module_type(self) -> str:
        return "excel_find_replace"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        find = context.resolve_value(config.get("find", ""))
        replace = context.resolve_value(config.get("replace", ""))
        match_entire = to_bool(config.get("matchEntire", False), context)
        if find is None or str(find) == "":
            return ModuleResult(success=False, error="查找内容不能为空")
        try:
            find_s = str(find)
            repl_s = str(replace if replace is not None else "")
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            count = 0
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    cv = str(cell.value)
                    if match_entire:
                        if cv == find_s:
                            cell.value = repl_s
                            count += 1
                    else:
                        if find_s in cv:
                            cell.value = cv.replace(find_s, repl_s)
                            count += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已替换 {count} 处「{find_s}」→「{repl_s}」", data={"count": count})
        except Exception as e:
            return ModuleResult(success=False, error=f"查找替换失败: {e}")


def _norm_color(c: str) -> str:
    """颜色归一化为 openpyxl 的 AARRGGBB / RRGGBB 十六进制（去 #，补全）。"""
    s = (c or "").strip().lstrip("#")
    if not s:
        return ""
    if len(s) == 6:
        return s.upper()
    if len(s) == 8:
        return s.upper()
    return s.upper()


@register_executor
class ExcelSetStyleExecutor(ModuleExecutor):
    """设置单元格/区域样式：字体、加粗、斜体、字号、字色、背景色、对齐、边框"""

    @property
    def module_type(self) -> str:
        return "excel_set_style"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1 或 A1:C3）")
        bold = to_bool(config.get("bold", False), context)
        italic = to_bool(config.get("italic", False), context)
        font_size = to_int(config.get("fontSize", 0), 0, context)
        font_name = context.resolve_value(config.get("fontName", "")) or ""
        font_color = _norm_color(context.resolve_value(config.get("fontColor", "")) or "")
        bg_color = _norm_color(context.resolve_value(config.get("bgColor", "")) or "")
        align_h = context.resolve_value(config.get("alignH", "")) or ""   # left/center/right
        align_v = context.resolve_value(config.get("alignV", "")) or ""   # top/center/bottom
        border = to_bool(config.get("border", False), context)
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cells = ws[cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            thin = Side(style="thin", color="000000")
            bd = Border(left=thin, right=thin, top=thin, bottom=thin)
            n = 0
            for row in cells:
                if not isinstance(row, tuple):
                    row = (row,)
                for cell in row:
                    # 字体（合并现有属性）
                    if bold or italic or font_size or font_name or font_color:
                        cur = cell.font
                        cell.font = Font(
                            name=font_name or cur.name,
                            size=font_size or cur.size,
                            bold=bold or cur.bold,
                            italic=italic or cur.italic,
                            color=("FF" + font_color if len(font_color) == 6 else font_color) if font_color else cur.color,
                        )
                    if bg_color:
                        fill_hex = "FF" + bg_color if len(bg_color) == 6 else bg_color
                        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
                    if align_h or align_v:
                        cell.alignment = Alignment(
                            horizontal=align_h or None,
                            vertical=align_v or None,
                            wrap_text=cell.alignment.wrap_text,
                        )
                    if border:
                        cell.border = bd
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置样式 {cell_range}（{n} 格）")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置样式失败: {e}")


@register_executor
class ExcelMergeCellsExecutor(ModuleExecutor):
    """合并 / 取消合并单元格"""

    @property
    def module_type(self) -> str:
        return "excel_merge_cells"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        unmerge = to_bool(config.get("unmerge", False), context)
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:C1）")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if unmerge:
                ws.unmerge_cells(cell_range)
                msg = f"已取消合并 {cell_range}"
            else:
                ws.merge_cells(cell_range)
                msg = f"已合并 {cell_range}"
            wb.save(path)
            return ModuleResult(success=True, message=msg)
        except Exception as e:
            return ModuleResult(success=False, error=f"合并单元格失败: {e}")


@register_executor
class ExcelSetSizeExecutor(ModuleExecutor):
    """设置列宽 / 行高"""

    @property
    def module_type(self) -> str:
        return "excel_set_size"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        target = context.resolve_value(config.get("target", "column")) or "column"  # column / row
        key = (context.resolve_value(config.get("key", "")) or "").strip()  # 列字母(A) 或 行号(1)
        size = to_int(config.get("size", 0), 0, context)
        if not key or size <= 0:
            return ModuleResult(success=False, error="请填写列字母/行号(key)与尺寸(size>0)")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if str(target) == "row":
                ws.row_dimensions[int(key)].height = size
                msg = f"已设置第 {key} 行行高 = {size}"
            else:
                ws.column_dimensions[key.upper()].width = size
                msg = f"已设置 {key.upper()} 列列宽 = {size}"
            wb.save(path)
            return ModuleResult(success=True, message=msg)
        except Exception as e:
            return ModuleResult(success=False, error=f"设置尺寸失败: {e}")


@register_executor
class ExcelFreezePanesExecutor(ModuleExecutor):
    """冻结窗格（冻结到指定单元格左上，如 A2 冻结首行）"""

    @property
    def module_type(self) -> str:
        return "excel_freeze_panes"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "A2")) or "A2").strip().upper()
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.freeze_panes = cell if cell.upper() not in ("", "NONE") else None
            wb.save(path)
            return ModuleResult(success=True, message=f"已冻结窗格至 {cell}")
        except Exception as e:
            return ModuleResult(success=False, error=f"冻结窗格失败: {e}")


@register_executor
class ExcelGetInfoExecutor(ModuleExecutor):
    """获取表格信息（最大行列数、维度），写入变量"""

    @property
    def module_type(self) -> str:
        return "excel_get_info"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            info = {
                "sheet": ws.title,
                "maxRow": ws.max_row,
                "maxColumn": ws.max_column,
                "dimensions": ws.calculate_dimension(),
                "sheets": list(wb.sheetnames),
            }
            if var:
                context.set_variable(var, info)
            return ModuleResult(success=True, message=f"{ws.title}: {info['maxRow']} 行 × {info['maxColumn']} 列", data=info)
        except Exception as e:
            return ModuleResult(success=False, error=f"获取表格信息失败: {e}")
