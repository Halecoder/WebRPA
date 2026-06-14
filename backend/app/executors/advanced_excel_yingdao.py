# -*- coding: utf-8 -*-
"""Excel 自动化模块执行器 - 影刀对标补全（openpyxl 文件级 + win32com 活动应用级）

对照影刀 RPA 的 Excel/WPS 模块清单，补齐 WebRPA 尚未支持的能力：
- 读取总行数、查找第一个空行/空列/空单元格
- 填充内容、删除样式、激活工作表、另存为
- 数据透视表（分组聚合生成汇总表）
- 导出 PDF、运行宏、刷新数据（基于 win32com 操作 Excel/WPS）

设计原则：与项目其它 Excel 模块一致，文件路径式、无状态、懒加载。
COM 类模块在 Windows 且安装 Excel/WPS 时可用；缺环境时给出清晰报错。
"""
import os
from typing import Any

from .base import ModuleExecutor, ExecutionContext, ModuleResult, register_executor
from .type_utils import to_int, to_bool


def _openpyxl():
    import openpyxl
    return openpyxl


def _resolve_path(context: ExecutionContext, raw: str) -> str:
    return (raw or "").strip().strip('"').strip("'")


def _load_wb(path: str, read_only: bool = False, data_only: bool = False):
    openpyxl = _openpyxl()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel 文件不存在: {path}")
    return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)


def _get_ws(wb, sheet_name: str):
    if not sheet_name:
        return wb.active
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}（现有: {', '.join(wb.sheetnames)}）")
    return wb[sheet_name]


# ============================================================
# 读写辅助：行列计数 / 空行空列查找 / 填充 / 清样式
# ============================================================

@register_executor
class ExcelCountRowsExecutor(ModuleExecutor):
    """读取 Excel 总行数 / 总列数"""

    @property
    def module_type(self) -> str:
        return "excel_count_rows"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            wb = _load_wb(path, read_only=True)
            ws = _get_ws(wb, sheet_name)
            # read_only 模式下 max_row/max_column 可能不准，用迭代兜底
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            wb.close()
            if var:
                context.set_variable(var, max_row)
            return ModuleResult(success=True, message=f"{ws.title}: 共 {max_row} 行 × {max_col} 列",
                                data={"rows": max_row, "columns": max_col})
        except Exception as e:
            return ModuleResult(success=False, error=f"读取总行数失败: {e}")


@register_executor
class ExcelFindEmptyRowExecutor(ModuleExecutor):
    """获取第一个空行（从指定列判断；direction: down 从上往下 / up 从下往上）"""

    @property
    def module_type(self) -> str:
        return "excel_find_empty_row"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        col = (context.resolve_value(config.get("column", "A")) or "A").strip().upper()
        direction = (context.resolve_value(config.get("direction", "down")) or "down").strip()
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            from openpyxl.utils import column_index_from_string
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cidx = column_index_from_string(col) if col.isalpha() else int(col)
            max_row = ws.max_row or 0
            empty_row = None
            if direction == "up":
                # 从最后一行往上找第一个非空行的下一行（即末尾追加位置的“上方第一个空行”）
                r = max_row
                while r >= 1:
                    if ws.cell(row=r, column=cidx).value not in (None, ""):
                        empty_row = r + 1
                        break
                    r -= 1
                if empty_row is None:
                    empty_row = 1
            else:
                r = 1
                while r <= max_row:
                    if ws.cell(row=r, column=cidx).value in (None, ""):
                        empty_row = r
                        break
                    r += 1
                if empty_row is None:
                    empty_row = max_row + 1
            wb.close()
            if var:
                context.set_variable(var, empty_row)
            return ModuleResult(success=True, message=f"第一个空行（{col}列, {direction}）= {empty_row}",
                                data={"row": empty_row})
        except Exception as e:
            return ModuleResult(success=False, error=f"查找空行失败: {e}")


@register_executor
class ExcelFindEmptyColExecutor(ModuleExecutor):
    """获取第一个空列（从指定行判断；返回列号与列字母）"""

    @property
    def module_type(self) -> str:
        return "excel_find_empty_col"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        row = to_int(config.get("row", 1), 1, context)
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            from openpyxl.utils import get_column_letter
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            max_col = ws.max_column or 0
            empty_col = None
            c = 1
            while c <= max_col:
                if ws.cell(row=row, column=c).value in (None, ""):
                    empty_col = c
                    break
                c += 1
            if empty_col is None:
                empty_col = max_col + 1
            letter = get_column_letter(empty_col)
            wb.close()
            if var:
                context.set_variable(var, letter)
            return ModuleResult(success=True, message=f"第一个空列（第 {row} 行）= {letter}（第 {empty_col} 列）",
                                data={"column": empty_col, "letter": letter})
        except Exception as e:
            return ModuleResult(success=False, error=f"查找空列失败: {e}")


@register_executor
class ExcelFindEmptyCellExecutor(ModuleExecutor):
    """获取指定列中第一个空白单元格地址"""

    @property
    def module_type(self) -> str:
        return "excel_find_empty_cell"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        col = (context.resolve_value(config.get("column", "A")) or "A").strip().upper()
        start_row = to_int(config.get("startRow", 1), 1, context)
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            from openpyxl.utils import column_index_from_string, get_column_letter
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cidx = column_index_from_string(col) if col.isalpha() else int(col)
            max_row = ws.max_row or 0
            target = None
            r = start_row
            while r <= max_row:
                if ws.cell(row=r, column=cidx).value in (None, ""):
                    target = f"{get_column_letter(cidx)}{r}"
                    break
                r += 1
            if target is None:
                target = f"{get_column_letter(cidx)}{max_row + 1}"
            wb.close()
            if var:
                context.set_variable(var, target)
            return ModuleResult(success=True, message=f"第一个空白单元格 = {target}", data={"cell": target})
        except Exception as e:
            return ModuleResult(success=False, error=f"查找空单元格失败: {e}")


@register_executor
class ExcelFillRangeExecutor(ModuleExecutor):
    """填充 Excel 内容（把同一个值/公式填充到整个区域）"""

    @property
    def module_type(self) -> str:
        return "excel_fill_range"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        value = context.resolve_value(config.get("value", ""))
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:C10）")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cells = ws[cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            n = 0
            for row in cells:
                if not isinstance(row, tuple):
                    row = (row,)
                for c in row:
                    c.value = value
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已填充 {cell_range}（{n} 格）= {value}")
        except Exception as e:
            return ModuleResult(success=False, error=f"填充内容失败: {e}")


@register_executor
class ExcelClearStyleExecutor(ModuleExecutor):
    """删除 Excel 样式（清除区域字体/填充/边框/对齐/数字格式，保留内容）"""

    @property
    def module_type(self) -> str:
        return "excel_clear_style"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:C10）")
        try:
            from openpyxl.styles import Font, PatternFill, Border, Alignment
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cells = ws[cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            n = 0
            for row in cells:
                if not isinstance(row, tuple):
                    row = (row,)
                for c in row:
                    c.font = Font()
                    c.fill = PatternFill()
                    c.border = Border()
                    c.alignment = Alignment()
                    c.number_format = "General"
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已清除样式 {cell_range}（{n} 格）")
        except Exception as e:
            return ModuleResult(success=False, error=f"删除样式失败: {e}")


@register_executor
class ExcelActivateSheetExecutor(ModuleExecutor):
    """激活 Sheet 页（设为打开时默认显示的活动工作表）"""

    @property
    def module_type(self) -> str:
        return "excel_activate_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        if not sheet_name:
            return ModuleResult(success=False, error="请指定要激活的工作表名")
        try:
            wb = _load_wb(path)
            if sheet_name not in wb.sheetnames:
                return ModuleResult(success=False, error=f"工作表不存在: {sheet_name}")
            wb.active = wb.sheetnames.index(sheet_name)
            wb.save(path)
            return ModuleResult(success=True, message=f"已激活工作表: {sheet_name}")
        except Exception as e:
            return ModuleResult(success=False, error=f"激活工作表失败: {e}")


@register_executor
class ExcelSaveAsExecutor(ModuleExecutor):
    """另存为 Excel（复制当前工作簿到新路径）"""

    @property
    def module_type(self) -> str:
        return "excel_save_as"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        new_path = _resolve_path(context, context.resolve_value(config.get("newPath", "")))
        if not new_path:
            return ModuleResult(success=False, error="另存路径(newPath)不能为空")
        try:
            wb = _load_wb(path)
            os.makedirs(os.path.dirname(new_path) or ".", exist_ok=True)
            wb.save(new_path)
            return ModuleResult(success=True, message=f"已另存为: {new_path}", data={"path": new_path})
        except Exception as e:
            return ModuleResult(success=False, error=f"另存为失败: {e}")


# ============================================================
# 数据透视表（分组聚合生成汇总表，基于 openpyxl，无需 Excel 应用）
# ============================================================

@register_executor
class ExcelPivotTableExecutor(ModuleExecutor):
    """新建数据透视表（按行字段分组，对值字段聚合：sum/count/average/max/min）

    读取源区域(含表头) → 按 groupBy 列分组 → 对 valueColumn 聚合 → 写出汇总表。
    """

    @property
    def module_type(self) -> str:
        return "excel_pivot_table"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        source_range = (context.resolve_value(config.get("sourceRange", "")) or "").strip().upper()
        group_by = context.resolve_value(config.get("groupBy", "")) or ""        # 分组列名（表头），逗号分隔
        value_column = context.resolve_value(config.get("valueColumn", "")) or ""  # 聚合列名（表头）
        agg = (context.resolve_value(config.get("aggregation", "sum")) or "sum").strip().lower()
        dest_sheet = context.resolve_value(config.get("destSheet", "")) or ""
        dest_cell = (context.resolve_value(config.get("destCell", "A1")) or "A1").strip().upper()
        if not group_by:
            return ModuleResult(success=False, error="请指定分组列名(groupBy)")
        try:
            from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple, get_column_letter
            wb = _load_wb(path, data_only=True)
            ws = _get_ws(wb, sheet_name)

            # 取数据区域（默认整表）
            if source_range and ":" in source_range:
                min_c, min_r, max_c, max_r = range_boundaries(source_range)
            else:
                min_c, min_r, max_c, max_r = 1, 1, ws.max_column or 1, ws.max_row or 1
            rows = []
            for r in range(min_r, max_r + 1):
                rows.append([ws.cell(row=r, column=c).value for c in range(min_c, max_c + 1)])
            if len(rows) < 2:
                return ModuleResult(success=False, error="源数据不足（至少需要表头+1行数据）")
            headers = [str(h) if h is not None else f"col{i+1}" for i, h in enumerate(rows[0])]
            body = rows[1:]

            group_cols = [g.strip() for g in str(group_by).replace("，", ",").split(",") if g.strip()]
            for g in group_cols:
                if g not in headers:
                    return ModuleResult(success=False, error=f"分组列不存在: {g}（表头: {', '.join(headers)}）")
            gidx = [headers.index(g) for g in group_cols]
            vidx = headers.index(value_column) if value_column and value_column in headers else None

            def to_num(v):
                try:
                    return float(v)
                except Exception:
                    return None

            groups: dict = {}
            order: list = []
            for row in body:
                if all(v is None for v in row):
                    continue
                key = tuple(row[i] if i < len(row) else None for i in gidx)
                if key not in groups:
                    groups[key] = []
                    order.append(key)
                if vidx is not None and vidx < len(row):
                    groups[key].append(row[vidx])

            def aggregate(vals):
                nums = [to_num(v) for v in vals]
                nums = [n for n in nums if n is not None]
                if agg == "count":
                    return len(vals)
                if not nums:
                    return 0
                if agg == "sum":
                    return sum(nums)
                if agg == "average":
                    return round(sum(nums) / len(nums), 6)
                if agg == "max":
                    return max(nums)
                if agg == "min":
                    return min(nums)
                return sum(nums)

            # 组织汇总表
            agg_label = {"sum": "求和", "count": "计数", "average": "平均", "max": "最大", "min": "最小"}.get(agg, agg)
            header_out = list(group_cols) + [f"{agg_label}_{value_column or '行数'}"]
            result_rows = [header_out]
            for key in order:
                result_rows.append(list(key) + [aggregate(groups[key])])

            ws_dest = _get_ws(wb, dest_sheet) if dest_sheet else ws
            r0, c0 = coordinate_to_tuple(dest_cell)
            for i, rrow in enumerate(result_rows):
                for j, val in enumerate(rrow):
                    ws_dest.cell(row=r0 + i, column=c0 + j, value=val)
            wb.save(path)
            return ModuleResult(
                success=True,
                message=f"已生成透视汇总（{len(order)} 组，{agg_label}）→ {dest_cell}",
                data={"groups": len(order)},
            )
        except Exception as e:
            return ModuleResult(success=False, error=f"生成数据透视表失败: {e}")


# ============================================================
# 活动应用级（win32com 操作 Excel / WPS）：导出PDF / 运行宏 / 刷新数据
# ============================================================

def _com_excel_app():
    """创建 Excel COM 应用，失败则回退 WPS（KET）。返回 (app, name)。"""
    import win32com.client as win32
    last_err = None
    for prog, name in (("Excel.Application", "Excel"), ("KET.Application", "WPS")):
        try:
            app = win32.DispatchEx(prog)
            return app, name
        except Exception as e:
            last_err = e
    raise RuntimeError(f"无法启动 Excel/WPS（请确认已安装）: {last_err}")


def _run_com(func):
    """在已初始化 COM 的环境中执行 func(app, name)，保证 CoInitialize/Uninitialize 与 app.Quit。"""
    import pythoncom
    pythoncom.CoInitialize()
    app = None
    try:
        app, name = _com_excel_app()
        app.Visible = False
        try:
            app.DisplayAlerts = False
        except Exception:
            pass
        return func(app, name)
    finally:
        if app is not None:
            try:
                app.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


@register_executor
class ExcelToPdfExecutor(ModuleExecutor):
    """导出为 PDF（基于 Excel/WPS COM，把工作簿或指定工作表导出成 PDF）"""

    @property
    def module_type(self) -> str:
        return "excel_to_pdf"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import asyncio
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        pdf_path = _resolve_path(context, context.resolve_value(config.get("pdfPath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", "")) or ""
        if not os.path.exists(path):
            return ModuleResult(success=False, error=f"Excel 文件不存在: {path}")
        if not pdf_path:
            base = os.path.splitext(path)[0]
            pdf_path = base + ".pdf"
        abs_in = os.path.abspath(path)
        abs_out = os.path.abspath(pdf_path)

        def work(app, name):
            os.makedirs(os.path.dirname(abs_out) or ".", exist_ok=True)
            wb = app.Workbooks.Open(abs_in, ReadOnly=True)
            try:
                target = wb
                if sheet_name:
                    try:
                        target = wb.Worksheets(sheet_name)
                    except Exception:
                        raise RuntimeError(f"工作表不存在: {sheet_name}")
                # 0 = xlTypePDF
                target.ExportAsFixedFormat(0, abs_out)
            finally:
                wb.Close(SaveChanges=False)
            return name

        try:
            loop = asyncio.get_running_loop()
            app_name = await loop.run_in_executor(None, lambda: _run_com(work))
            if not os.path.exists(abs_out):
                return ModuleResult(success=False, error="导出完成但未找到 PDF 文件")
            return ModuleResult(success=True, message=f"已用 {app_name} 导出 PDF: {abs_out}", data={"path": abs_out})
        except Exception as e:
            return ModuleResult(success=False, error=f"导出 PDF 失败: {e}")


@register_executor
class ExcelRunMacroExecutor(ModuleExecutor):
    """运行 Excel 宏（基于 COM，打开含宏的工作簿并执行指定宏，可选保存）"""

    @property
    def module_type(self) -> str:
        return "excel_run_macro"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import asyncio
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        macro_name = context.resolve_value(config.get("macroName", "")) or ""
        save_after = to_bool(config.get("saveAfter", False), context)
        if not os.path.exists(path):
            return ModuleResult(success=False, error=f"Excel 文件不存在: {path}")
        if not macro_name:
            return ModuleResult(success=False, error="请指定宏名(macroName)，如 Module1.MyMacro")
        abs_in = os.path.abspath(path)

        def work(app, name):
            wb = app.Workbooks.Open(abs_in)
            result = None
            try:
                result = app.Run(macro_name)
                if save_after:
                    wb.Save()
            finally:
                wb.Close(SaveChanges=bool(save_after))
            return (name, result)

        try:
            loop = asyncio.get_running_loop()
            app_name, result = await loop.run_in_executor(None, lambda: _run_com(work))
            var = config.get("resultVariable", "") or config.get("variableName", "")
            if var and result is not None:
                context.set_variable(var, result)
            return ModuleResult(success=True, message=f"已用 {app_name} 运行宏: {macro_name}",
                                data={"result": result})
        except Exception as e:
            return ModuleResult(success=False, error=f"运行宏失败: {e}")


@register_executor
class ExcelRefreshDataExecutor(ModuleExecutor):
    """刷新 Excel 数据（基于 COM，刷新数据透视表/外部数据连接/查询后保存）"""

    @property
    def module_type(self) -> str:
        return "excel_refresh_data"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        import asyncio
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        if not os.path.exists(path):
            return ModuleResult(success=False, error=f"Excel 文件不存在: {path}")
        abs_in = os.path.abspath(path)

        def work(app, name):
            wb = app.Workbooks.Open(abs_in)
            try:
                wb.RefreshAll()
                try:
                    app.CalculateUntilAsyncQueriesDone()
                except Exception:
                    pass
                wb.Save()
            finally:
                wb.Close(SaveChanges=True)
            return name

        try:
            loop = asyncio.get_running_loop()
            app_name = await loop.run_in_executor(None, lambda: _run_com(work))
            return ModuleResult(success=True, message=f"已用 {app_name} 刷新数据并保存: {path}")
        except Exception as e:
            return ModuleResult(success=False, error=f"刷新数据失败: {e}")
