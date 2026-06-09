# -*- coding: utf-8 -*-
"""Excel 自动化模块执行器 - 高级增强版（基于 openpyxl）

在 advanced_openpyxl.py 的基础上，补齐 openpyxl 几乎所有 Excel 自动化能力：
工作表复制/移动/标签色、区域复制、数字格式、超链接、批注、图片、图表、
数据验证(下拉)、条件格式、隐藏行列、自动筛选、排序、去重、字典读写、
清空整表、工作表保护、CSV 互转、页面/打印设置、读取公式/计算值、视图缩放等。

设计原则（与项目其它模块一致）：
- 文件路径式、无状态：每个模块打开文件 → 操作 → 保存
- openpyxl 懒加载（函数内 import）
- 所有配置值都过 context.resolve_value，支持 {变量名}
"""
import os
import json
from typing import Any

from .base import ModuleExecutor, ExecutionContext, ModuleResult, register_executor
from .type_utils import to_int, to_bool, to_float


def _openpyxl():
    import openpyxl
    return openpyxl


def _resolve_path(context: ExecutionContext, raw: str) -> str:
    return (raw or "").strip().strip('"').strip("'")


def _load_wb(path: str, read_only: bool = False, data_only: bool = False):
    openpyxl = _openpyxl()
    if not os.path.exists(path):
        raise FileNotFoundError(f"Excel 文件不存在: {path}")
    return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)


def _get_ws(wb, sheet_name: str):
    if not sheet_name:
        return wb.active
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表不存在: {sheet_name}（现有: {', '.join(wb.sheetnames)}）")
    return wb[sheet_name]


def _parse_2d(raw: Any) -> list:
    if isinstance(raw, (list, tuple)):
        data = list(raw)
    else:
        s = str(raw or "").strip()
        if not s:
            return []
        data = json.loads(s)
    if not isinstance(data, list):
        return [[data]]
    if data and not isinstance(data[0], (list, tuple)):
        return [data]
    return [list(r) for r in data]


def _parse_1d(raw: Any) -> list:
    if isinstance(raw, (list, tuple)):
        return list(raw)
    s = str(raw or "").strip()
    if not s:
        return []
    try:
        v = json.loads(s)
    except Exception:
        return [x.strip() for x in s.split(",")]
    if isinstance(v, list):
        return v
    return [v]


def _norm_color(c: str) -> str:
    """颜色归一化为 openpyxl 的十六进制（去 #，6 位补成 8 位 AARRGGBB）。"""
    s = (c or "").strip().lstrip("#").upper()
    if not s:
        return ""
    if len(s) == 6:
        return "FF" + s
    return s


# ============================================================
# 工作表 / 区域 管理
# ============================================================

@register_executor
class ExcelCopySheetExecutor(ModuleExecutor):
    """复制工作表（在同一工作簿内复制一份）"""

    @property
    def module_type(self) -> str:
        return "excel_copy_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        new_name = context.resolve_value(config.get("newName", "")) or ""
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            target = wb.copy_worksheet(ws)
            if new_name:
                if new_name in wb.sheetnames and new_name != target.title:
                    return ModuleResult(success=False, error=f"工作表已存在: {new_name}")
                target.title = new_name
            wb.save(path)
            return ModuleResult(success=True, message=f"已复制工作表 {ws.title} → {target.title}",
                                data={"sheets": wb.sheetnames, "newSheet": target.title})
        except Exception as e:
            return ModuleResult(success=False, error=f"复制工作表失败: {e}")


@register_executor
class ExcelMoveSheetExecutor(ModuleExecutor):
    """移动工作表顺序（offset 为正向右移，负向左移）"""

    @property
    def module_type(self) -> str:
        return "excel_move_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        offset = to_int(config.get("offset", 0), 0, context)
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            wb.move_sheet(ws, offset=offset)
            wb.save(path)
            return ModuleResult(success=True, message=f"已移动工作表 {ws.title}（offset={offset}）",
                                data={"sheets": wb.sheetnames})
        except Exception as e:
            return ModuleResult(success=False, error=f"移动工作表失败: {e}")


@register_executor
class ExcelSetTabColorExecutor(ModuleExecutor):
    """设置工作表标签颜色（十六进制，如 FF0000 红色；留空清除）"""

    @property
    def module_type(self) -> str:
        return "excel_set_tab_color"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        color = _norm_color(context.resolve_value(config.get("color", "")) or "")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.sheet_properties.tabColor = color or None
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置 {ws.title} 标签颜色 = {color or '无'}")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置标签颜色失败: {e}")


@register_executor
class ExcelClearSheetExecutor(ModuleExecutor):
    """清空整个工作表内容（删除所有行后保留空表）"""

    @property
    def module_type(self) -> str:
        return "excel_clear_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            max_row = ws.max_row
            if max_row and max_row > 0:
                ws.delete_rows(1, max_row)
            wb.save(path)
            return ModuleResult(success=True, message=f"已清空工作表 {ws.title}")
        except Exception as e:
            return ModuleResult(success=False, error=f"清空工作表失败: {e}")


@register_executor
class ExcelCopyRangeExecutor(ModuleExecutor):
    """复制区域内容到另一位置（同表或跨表，仅复制值）"""

    @property
    def module_type(self) -> str:
        return "excel_copy_range"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        src_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        dest_sheet = context.resolve_value(config.get("destSheet", "")) or ""
        dest_cell = (context.resolve_value(config.get("destCell", "")) or "").strip().upper()
        if not src_range or not dest_cell:
            return ModuleResult(success=False, error="请填写源区域(range)与目标起始单元格(destCell)")
        try:
            from openpyxl.utils.cell import coordinate_to_tuple
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws_dest = _get_ws(wb, dest_sheet) if dest_sheet else ws
            cells = ws[src_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            r0, c0 = coordinate_to_tuple(dest_cell)
            n = 0
            for i, row in enumerate(cells):
                if not isinstance(row, tuple):
                    row = (row,)
                for j, cell in enumerate(row):
                    ws_dest.cell(row=r0 + i, column=c0 + j, value=cell.value)
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已复制区域 {src_range} → {dest_cell}（{n} 格）")
        except Exception as e:
            return ModuleResult(success=False, error=f"复制区域失败: {e}")


# ============================================================
# 格式 / 批注 / 超链接
# ============================================================

# 数字格式预设
_NUMBER_FORMATS = {
    "general": "General",
    "integer": "0",
    "decimal2": "0.00",
    "thousands": "#,##0",
    "thousands2": "#,##0.00",
    "percent": "0%",
    "percent2": "0.00%",
    "currency_cny": "¥#,##0.00",
    "currency_usd": "$#,##0.00",
    "date": "yyyy-mm-dd",
    "datetime": "yyyy-mm-dd hh:mm:ss",
    "time": "hh:mm:ss",
    "text": "@",
    "scientific": "0.00E+00",
}


@register_executor
class ExcelNumberFormatExecutor(ModuleExecutor):
    """设置单元格/区域数字格式（日期、货币、百分比、千分位等）"""

    @property
    def module_type(self) -> str:
        return "excel_number_format"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        preset = (context.resolve_value(config.get("preset", "")) or "").strip()
        custom = context.resolve_value(config.get("customFormat", "")) or ""
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1 或 A1:C10）")
        fmt = custom.strip() if custom and custom.strip() else _NUMBER_FORMATS.get(preset, preset)
        if not fmt:
            return ModuleResult(success=False, error="请选择格式预设或填写自定义格式")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cells = ws[cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            n = 0
            for row in cells:
                if not isinstance(row, tuple):
                    row = (row,)
                for cell in row:
                    cell.number_format = fmt
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置数字格式 {cell_range} = {fmt}（{n} 格）")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置数字格式失败: {e}")


@register_executor
class ExcelSetHyperlinkExecutor(ModuleExecutor):
    """设置单元格超链接（网址 / 文件路径 / 内部跳转）"""

    @property
    def module_type(self) -> str:
        return "excel_set_hyperlink"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "")) or "").strip().upper()
        link = context.resolve_value(config.get("link", "")) or ""
        display = context.resolve_value(config.get("display", "")) or ""
        if not cell or not link:
            return ModuleResult(success=False, error="请填写单元格(cell)与链接(link)")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            c = ws[cell]
            c.hyperlink = link
            c.value = display or (c.value if c.value is not None else link)
            from openpyxl.styles import Font
            c.font = Font(color="0563C1", underline="single")
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置超链接 {cell} → {link}")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置超链接失败: {e}")


@register_executor
class ExcelSetCommentExecutor(ModuleExecutor):
    """设置/清除单元格批注"""

    @property
    def module_type(self) -> str:
        return "excel_set_comment"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "")) or "").strip().upper()
        text = context.resolve_value(config.get("text", ""))
        author = context.resolve_value(config.get("author", "")) or "WebRPA"
        if not cell:
            return ModuleResult(success=False, error="单元格地址不能为空")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            c = ws[cell]
            if text is None or str(text) == "":
                c.comment = None
                msg = f"已清除批注 {cell}"
            else:
                from openpyxl.comments import Comment
                c.comment = Comment(str(text), str(author))
                msg = f"已设置批注 {cell}"
            wb.save(path)
            return ModuleResult(success=True, message=msg)
        except Exception as e:
            return ModuleResult(success=False, error=f"设置批注失败: {e}")


@register_executor
class ExcelSetBorderExecutor(ModuleExecutor):
    """设置区域边框（样式 thin/medium/thick/dashed/dotted/double，颜色，范围 all/outline）"""

    @property
    def module_type(self) -> str:
        return "excel_set_border"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        style = (context.resolve_value(config.get("style", "thin")) or "thin").strip()
        color = _norm_color(context.resolve_value(config.get("color", "000000")) or "000000")
        scope = (context.resolve_value(config.get("scope", "all")) or "all").strip()  # all / outline
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:C5）")
        try:
            from openpyxl.styles import Border, Side
            side = Side(style=style, color=color)
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cells = ws[cell_range]
            if not isinstance(cells, tuple):
                cells = ((cells,),)
            rows = [r if isinstance(r, tuple) else (r,) for r in cells]
            nrows = len(rows)
            ncols = len(rows[0]) if rows else 0
            n = 0
            for i, row in enumerate(rows):
                for j, cell in enumerate(row):
                    if scope == "outline":
                        cell.border = Border(
                            left=side if j == 0 else cell.border.left,
                            right=side if j == ncols - 1 else cell.border.right,
                            top=side if i == 0 else cell.border.top,
                            bottom=side if i == nrows - 1 else cell.border.bottom,
                        )
                    else:
                        cell.border = Border(left=side, right=side, top=side, bottom=side)
                    n += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置边框 {cell_range}（{n} 格，{scope}）")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置边框失败: {e}")


# ============================================================
# 图片 / 图表 / 数据验证 / 条件格式
# ============================================================

@register_executor
class ExcelAddImageExecutor(ModuleExecutor):
    """在工作表插入图片（指定锚点单元格）"""

    @property
    def module_type(self) -> str:
        return "excel_add_image"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        img_path = _resolve_path(context, context.resolve_value(config.get("imagePath", "")))
        anchor = (context.resolve_value(config.get("anchor", "A1")) or "A1").strip().upper()
        width = to_int(config.get("width", 0), 0, context)
        height = to_int(config.get("height", 0), 0, context)
        if not img_path or not os.path.exists(img_path):
            return ModuleResult(success=False, error=f"图片不存在: {img_path}")
        try:
            from openpyxl.drawing.image import Image as XLImage
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            img = XLImage(img_path)
            if width > 0:
                img.width = width
            if height > 0:
                img.height = height
            ws.add_image(img, anchor)
            wb.save(path)
            return ModuleResult(success=True, message=f"已插入图片到 {anchor}")
        except Exception as e:
            return ModuleResult(success=False, error=f"插入图片失败: {e}")


@register_executor
class ExcelAddChartExecutor(ModuleExecutor):
    """插入图表（柱状/折线/饼图/条形/面积/散点），基于数据区域"""

    @property
    def module_type(self) -> str:
        return "excel_add_chart"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        chart_type = (context.resolve_value(config.get("chartType", "bar")) or "bar").strip().lower()
        data_range = (context.resolve_value(config.get("dataRange", "")) or "").strip().upper()
        cats_range = (context.resolve_value(config.get("catsRange", "")) or "").strip().upper()
        anchor = (context.resolve_value(config.get("anchor", "")) or "").strip().upper()
        title = context.resolve_value(config.get("title", "")) or ""
        titles_from_data = to_bool(config.get("titlesFromData", True), context)
        if not data_range:
            return ModuleResult(success=False, error="数据区域(dataRange)不能为空，如 B1:B10")
        try:
            from openpyxl.chart import BarChart, LineChart, PieChart, AreaChart, ScatterChart, Reference, Series
            chart_map = {
                "bar": BarChart, "column": BarChart, "line": LineChart,
                "pie": PieChart, "area": AreaChart, "scatter": ScatterChart,
            }
            cls = chart_map.get(chart_type, BarChart)
            chart = cls()
            if chart_type == "bar":
                chart.type = "bar"
            elif chart_type == "column":
                chart.type = "col"
            if title:
                chart.title = title
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            from openpyxl.utils.cell import range_boundaries
            min_c, min_r, max_c, max_r = range_boundaries(data_range)
            data_ref = Reference(ws, min_col=min_c, min_row=min_r, max_col=max_c, max_row=max_r)
            chart.add_data(data_ref, titles_from_data=titles_from_data)
            if cats_range:
                cmin_c, cmin_r, cmax_c, cmax_r = range_boundaries(cats_range)
                cats = Reference(ws, min_col=cmin_c, min_row=cmin_r, max_col=cmax_c, max_row=cmax_r)
                chart.set_categories(cats)
            target = anchor or "H2"
            ws.add_chart(chart, target)
            wb.save(path)
            return ModuleResult(success=True, message=f"已插入 {chart_type} 图表到 {target}")
        except Exception as e:
            return ModuleResult(success=False, error=f"插入图表失败: {e}")


@register_executor
class ExcelDataValidationExecutor(ModuleExecutor):
    """添加数据验证：下拉列表(list)、整数(whole)、小数(decimal)、长度(textLength)"""

    @property
    def module_type(self) -> str:
        return "excel_data_validation"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        vtype = (context.resolve_value(config.get("validationType", "list")) or "list").strip()
        options_raw = context.resolve_value(config.get("options", ""))
        operator = (context.resolve_value(config.get("operator", "between")) or "between").strip()
        formula1 = context.resolve_value(config.get("formula1", "")) or ""
        formula2 = context.resolve_value(config.get("formula2", "")) or ""
        prompt = context.resolve_value(config.get("prompt", "")) or ""
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A2:A100）")
        try:
            from openpyxl.worksheet.datavalidation import DataValidation
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if vtype == "list":
                opts = _parse_1d(options_raw)
                if not opts:
                    return ModuleResult(success=False, error="下拉列表需要提供选项(options)，如 [\"是\",\"否\"]")
                formula = '"' + ",".join(str(o) for o in opts) + '"'
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            else:
                dv = DataValidation(type=vtype, operator=operator,
                                    formula1=str(formula1) or None,
                                    formula2=str(formula2) or None, allow_blank=True)
            if prompt:
                dv.prompt = prompt
                dv.promptTitle = "提示"
            ws.add_data_validation(dv)
            dv.add(cell_range)
            wb.save(path)
            return ModuleResult(success=True, message=f"已为 {cell_range} 添加数据验证（{vtype}）")
        except Exception as e:
            return ModuleResult(success=False, error=f"添加数据验证失败: {e}")


@register_executor
class ExcelConditionalFormatExecutor(ModuleExecutor):
    """条件格式：单元格规则(cellIs)、包含文本(containsText)、色阶(colorScale)、数据条(dataBar)"""

    @property
    def module_type(self) -> str:
        return "excel_conditional_format"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        rule_type = (context.resolve_value(config.get("ruleType", "cellIs")) or "cellIs").strip()
        operator = (context.resolve_value(config.get("operator", "greaterThan")) or "greaterThan").strip()
        value1 = context.resolve_value(config.get("value1", "")) or ""
        value2 = context.resolve_value(config.get("value2", "")) or ""
        text = context.resolve_value(config.get("text", "")) or ""
        bg_color = _norm_color(context.resolve_value(config.get("bgColor", "FFFF00")) or "FFFF00")
        if not cell_range:
            return ModuleResult(success=False, error="区域不能为空（如 A1:A20）")
        try:
            from openpyxl.styles import PatternFill, Font
            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule, Rule
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
            if rule_type == "colorScale":
                rule = ColorScaleRule(start_type="min", start_color="FFF8696B",
                                      mid_type="percentile", mid_value=50, mid_color="FFFFEB84",
                                      end_type="max", end_color="FF63BE7B")
                ws.conditional_formatting.add(cell_range, rule)
            elif rule_type == "dataBar":
                rule = DataBarRule(start_type="min", end_type="max", color=bg_color.lstrip("F")[-6:] or "638EC6")
                ws.conditional_formatting.add(cell_range, rule)
            elif rule_type == "containsText":
                from openpyxl.formatting.rule import Rule
                from openpyxl.styles.differential import DifferentialStyle
                dxf = DifferentialStyle(fill=fill)
                rule = Rule(type="containsText", operator="containsText", text=str(text), dxf=dxf)
                rule.formula = [f'NOT(ISERROR(SEARCH("{text}",{cell_range.split(":")[0]})))']
                ws.conditional_formatting.add(cell_range, rule)
            else:  # cellIs
                formulas = [str(value1)]
                if operator in ("between", "notBetween"):
                    formulas.append(str(value2))
                rule = CellIsRule(operator=operator, formula=formulas, fill=fill)
                ws.conditional_formatting.add(cell_range, rule)
            wb.save(path)
            return ModuleResult(success=True, message=f"已为 {cell_range} 添加条件格式（{rule_type}）")
        except Exception as e:
            return ModuleResult(success=False, error=f"添加条件格式失败: {e}")


# ============================================================
# 行列显隐 / 筛选 / 排序 / 去重 / 字典读写
# ============================================================

@register_executor
class ExcelHideExecutor(ModuleExecutor):
    """隐藏 / 显示 行或列"""

    @property
    def module_type(self) -> str:
        return "excel_hide"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        target = (context.resolve_value(config.get("target", "column")) or "column").strip()  # column/row
        key = (context.resolve_value(config.get("key", "")) or "").strip()  # 列字母(A/A:C) 或 行号(1/1:3)
        hidden = to_bool(config.get("hidden", True), context)
        if not key:
            return ModuleResult(success=False, error="请填写列字母或行号(key)，如 A、A:C、1、1:5")
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            def _range(k):
                k = k.strip()
                if ":" in k:
                    a, b = k.split(":", 1)
                    return a.strip(), b.strip()
                return k, k
            a, b = _range(key)
            if target == "row":
                for i in range(int(a), int(b) + 1):
                    ws.row_dimensions[i].hidden = hidden
                msg = f"已{'隐藏' if hidden else '显示'}第 {key} 行"
            else:
                ws.column_dimensions.group(a.upper(), b.upper(), hidden=hidden)
                msg = f"已{'隐藏' if hidden else '显示'} {key} 列"
            wb.save(path)
            return ModuleResult(success=True, message=msg)
        except Exception as e:
            return ModuleResult(success=False, error=f"显隐操作失败: {e}")


@register_executor
class ExcelAutoFilterExecutor(ModuleExecutor):
    """设置自动筛选（对指定区域开启筛选；留空区域则清除）"""

    @property
    def module_type(self) -> str:
        return "excel_auto_filter"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if cell_range:
                ws.auto_filter.ref = cell_range
                msg = f"已对 {cell_range} 开启自动筛选"
            else:
                ws.auto_filter.ref = None
                msg = "已清除自动筛选"
            wb.save(path)
            return ModuleResult(success=True, message=msg)
        except Exception as e:
            return ModuleResult(success=False, error=f"设置自动筛选失败: {e}")


@register_executor
class ExcelSortRangeExecutor(ModuleExecutor):
    """对数据区域按某列排序（真正重排数据；可保留表头）"""

    @property
    def module_type(self) -> str:
        return "excel_sort_range"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        sort_col = to_int(config.get("sortColumn", 1), 1, context)  # 区域内第几列（从1开始）
        descending = to_bool(config.get("descending", False), context)
        has_header = to_bool(config.get("hasHeader", True), context)
        if not cell_range or ":" not in cell_range:
            return ModuleResult(success=False, error="区域不能为空且需为范围（如 A1:D100）")
        try:
            from openpyxl.utils.cell import range_boundaries, get_column_letter
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            min_c, min_r, max_c, max_r = range_boundaries(cell_range)
            rows = []
            for r in range(min_r, max_r + 1):
                rows.append([ws.cell(row=r, column=c).value for c in range(min_c, max_c + 1)])
            header = rows[:1] if has_header else []
            body = rows[1:] if has_header else rows
            idx = max(0, sort_col - 1)

            def _key(row):
                v = row[idx] if idx < len(row) else None
                return (v is None, isinstance(v, str), v if v is not None else "")
            body.sort(key=_key, reverse=descending)
            new_rows = header + body
            for i, row in enumerate(new_rows):
                for j, val in enumerate(row):
                    ws.cell(row=min_r + i, column=min_c + j, value=val)
            wb.save(path)
            return ModuleResult(success=True, message=f"已按第 {sort_col} 列{'降序' if descending else '升序'}排序 {len(body)} 行")
        except Exception as e:
            return ModuleResult(success=False, error=f"区域排序失败: {e}")


@register_executor
class ExcelRemoveDuplicatesExecutor(ModuleExecutor):
    """删除重复行（按整行或指定列判重，保留首次出现）"""

    @property
    def module_type(self) -> str:
        return "excel_remove_duplicates"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell_range = (context.resolve_value(config.get("range", "")) or "").strip().upper()
        key_cols_raw = context.resolve_value(config.get("keyColumns", ""))  # 如 "1,2" 区域内列号
        has_header = to_bool(config.get("hasHeader", True), context)
        try:
            from openpyxl.utils.cell import range_boundaries
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if cell_range and ":" in cell_range:
                min_c, min_r, max_c, max_r = range_boundaries(cell_range)
            else:
                min_c, min_r, max_c, max_r = 1, 1, ws.max_column, ws.max_row
            key_cols = [int(x) - 1 for x in str(key_cols_raw).replace("，", ",").split(",") if str(x).strip().isdigit()]
            rows = []
            for r in range(min_r, max_r + 1):
                rows.append([ws.cell(row=r, column=c).value for c in range(min_c, max_c + 1)])
            header = rows[:1] if has_header else []
            body = rows[1:] if has_header else rows
            seen = set()
            unique = []
            removed = 0
            for row in body:
                if key_cols:
                    sig = tuple(row[i] if i < len(row) else None for i in key_cols)
                else:
                    sig = tuple(row)
                if sig in seen:
                    removed += 1
                    continue
                seen.add(sig)
                unique.append(row)
            new_rows = header + unique
            # 清空原区域再写回
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    ws.cell(row=r, column=c, value=None)
            for i, row in enumerate(new_rows):
                for j, val in enumerate(row):
                    ws.cell(row=min_r + i, column=min_c + j, value=val)
            wb.save(path)
            return ModuleResult(success=True, message=f"已删除 {removed} 行重复，保留 {len(unique)} 行", data={"removed": removed})
        except Exception as e:
            return ModuleResult(success=False, error=f"删除重复行失败: {e}")


@register_executor
class ExcelWriteDictsExecutor(ModuleExecutor):
    """写入字典数组（自动生成表头，第一行为列名）"""

    @property
    def module_type(self) -> str:
        return "excel_write_dicts"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        data_raw = context.resolve_value(config.get("data", ""))
        write_header = to_bool(config.get("writeHeader", True), context)
        start_cell = (context.resolve_value(config.get("startCell", "A1")) or "A1").strip().upper()
        try:
            if isinstance(data_raw, (list, tuple)):
                rows = list(data_raw)
            else:
                rows = json.loads(str(data_raw or "[]"))
            if not isinstance(rows, list) or not rows:
                return ModuleResult(success=False, error="数据需为非空字典数组，如 [{\"姓名\":\"张三\"}]")
            if not isinstance(rows[0], dict):
                return ModuleResult(success=False, error="数组元素必须是字典对象")
            # 收集所有键作为列（保持首次出现顺序）
            keys = []
            for d in rows:
                for k in d.keys():
                    if k not in keys:
                        keys.append(k)
            from openpyxl.utils.cell import coordinate_to_tuple
            r0, c0 = coordinate_to_tuple(start_cell)
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            cur = r0
            if write_header:
                for j, k in enumerate(keys):
                    ws.cell(row=cur, column=c0 + j, value=k)
                cur += 1
            for d in rows:
                for j, k in enumerate(keys):
                    ws.cell(row=cur, column=c0 + j, value=d.get(k))
                cur += 1
            wb.save(path)
            return ModuleResult(success=True, message=f"已写入 {len(rows)} 条记录（{len(keys)} 列）")
        except Exception as e:
            return ModuleResult(success=False, error=f"写入字典数组失败: {e}")


@register_executor
class ExcelReadDictsExecutor(ModuleExecutor):
    """读取为字典数组（首行作为列名），存入变量"""

    @property
    def module_type(self) -> str:
        return "excel_read_dicts"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        header_row = to_int(config.get("headerRow", 1), 1, context)
        var = config.get("resultVariable", "") or config.get("variableName", "")
        try:
            wb = _load_wb(path, read_only=True, data_only=True)
            ws = _get_ws(wb, sheet_name)
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()
            if not all_rows or header_row > len(all_rows):
                result = []
            else:
                headers = [str(h) if h is not None else f"col{i+1}" for i, h in enumerate(all_rows[header_row - 1])]
                result = []
                for row in all_rows[header_row:]:
                    if all(v is None for v in row):
                        continue
                    result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
            if var:
                context.set_variable(var, result)
            return ModuleResult(success=True, message=f"已读取 {len(result)} 条记录", data={"data": result})
        except Exception as e:
            return ModuleResult(success=False, error=f"读取字典数组失败: {e}")


# ============================================================
# CSV 互转 / 保护 / 页面设置 / 公式 / 视图
# ============================================================

@register_executor
class ExcelToCsvExecutor(ModuleExecutor):
    """导出工作表为 CSV 文件"""

    @property
    def module_type(self) -> str:
        return "excel_to_csv"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        csv_path = _resolve_path(context, context.resolve_value(config.get("csvPath", "")))
        encoding = (context.resolve_value(config.get("encoding", "utf-8-sig")) or "utf-8-sig").strip()
        delimiter = context.resolve_value(config.get("delimiter", ",")) or ","
        if not csv_path:
            return ModuleResult(success=False, error="CSV 输出路径(csvPath)不能为空")
        try:
            import csv as _csv
            wb = _load_wb(path, read_only=True, data_only=True)
            ws = _get_ws(wb, sheet_name)
            os.makedirs(os.path.dirname(csv_path) or ".", exist_ok=True)
            n = 0
            with open(csv_path, "w", newline="", encoding=encoding) as f:
                writer = _csv.writer(f, delimiter=delimiter)
                for row in ws.iter_rows(values_only=True):
                    writer.writerow(["" if v is None else v for v in row])
                    n += 1
            wb.close()
            return ModuleResult(success=True, message=f"已导出 {n} 行到 CSV: {csv_path}", data={"path": csv_path, "rows": n})
        except Exception as e:
            return ModuleResult(success=False, error=f"导出 CSV 失败: {e}")


@register_executor
class ExcelFromCsvExecutor(ModuleExecutor):
    """CSV 文件转为 Excel（.xlsx）"""

    @property
    def module_type(self) -> str:
        return "excel_from_csv"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        csv_path = _resolve_path(context, context.resolve_value(config.get("csvPath", "")))
        out_path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        encoding = (context.resolve_value(config.get("encoding", "utf-8-sig")) or "utf-8-sig").strip()
        delimiter = context.resolve_value(config.get("delimiter", ",")) or ","
        sheet_name = context.resolve_value(config.get("sheetName", "")) or "Sheet1"
        if not csv_path or not os.path.exists(csv_path):
            return ModuleResult(success=False, error=f"CSV 文件不存在: {csv_path}")
        if not out_path:
            return ModuleResult(success=False, error="输出 Excel 路径(filePath)不能为空")
        try:
            import csv as _csv
            openpyxl = _openpyxl()
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = sheet_name
            n = 0
            with open(csv_path, "r", encoding=encoding, newline="") as f:
                reader = _csv.reader(f, delimiter=delimiter)
                for row in reader:
                    ws.append(row)
                    n += 1
            os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
            wb.save(out_path)
            return ModuleResult(success=True, message=f"已将 CSV 转为 Excel（{n} 行）: {out_path}", data={"path": out_path, "rows": n})
        except Exception as e:
            return ModuleResult(success=False, error=f"CSV 转 Excel 失败: {e}")


@register_executor
class ExcelProtectSheetExecutor(ModuleExecutor):
    """保护 / 取消保护工作表（可设密码）"""

    @property
    def module_type(self) -> str:
        return "excel_protect_sheet"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        protect = to_bool(config.get("protect", True), context)
        password = context.resolve_value(config.get("password", "")) or ""
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if protect:
                ws.protection.sheet = True
                if password:
                    ws.protection.password = password
                msg = f"已保护工作表 {ws.title}"
            else:
                ws.protection.sheet = False
                msg = f"已取消保护工作表 {ws.title}"
            wb.save(path)
            return ModuleResult(success=True, message=msg)
        except Exception as e:
            return ModuleResult(success=False, error=f"工作表保护操作失败: {e}")


@register_executor
class ExcelPageSetupExecutor(ModuleExecutor):
    """页面/打印设置：纸张(A4/A3/Letter)、方向(portrait/landscape)、缩放适配"""

    @property
    def module_type(self) -> str:
        return "excel_page_setup"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        orientation = (context.resolve_value(config.get("orientation", "")) or "").strip()
        paper = (context.resolve_value(config.get("paperSize", "")) or "").strip().upper()
        fit_width = to_int(config.get("fitToWidth", 0), 0, context)
        fit_height = to_int(config.get("fitToHeight", 0), 0, context)
        print_area = (context.resolve_value(config.get("printArea", "")) or "").strip().upper()
        try:
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            if orientation in ("portrait", "landscape"):
                ws.page_setup.orientation = orientation
            paper_map = {"A4": "9", "A3": "8", "LETTER": "1", "A5": "11"}
            if paper in paper_map:
                ws.page_setup.paperSize = paper_map[paper]
            if fit_width or fit_height:
                ws.sheet_properties.pageSetUpPr.fitToPage = True
                ws.page_setup.fitToWidth = fit_width or 1
                ws.page_setup.fitToHeight = fit_height or 1
            if print_area:
                ws.print_area = print_area
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置 {ws.title} 页面/打印参数")
        except Exception as e:
            return ModuleResult(success=False, error=f"页面设置失败: {e}")


@register_executor
class ExcelReadFormulaExecutor(ModuleExecutor):
    """读取单元格的公式文本或缓存计算值（mode: formula 公式 / value 计算值）"""

    @property
    def module_type(self) -> str:
        return "excel_read_formula"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        cell = (context.resolve_value(config.get("cell", "")) or "").strip().upper()
        mode = (context.resolve_value(config.get("mode", "value")) or "value").strip()
        var = config.get("resultVariable", "") or config.get("variableName", "")
        if not cell:
            return ModuleResult(success=False, error="单元格地址不能为空")
        try:
            data_only = (mode == "value")
            wb = _load_wb(path, read_only=True, data_only=data_only)
            ws = _get_ws(wb, sheet_name)
            value = ws[cell].value
            wb.close()
            if var:
                context.set_variable(var, value)
            tip = "（计算值需 Excel 打开保存过才有缓存）" if data_only and value is None else ""
            return ModuleResult(success=True, message=f"{cell} = {value} {tip}".strip(), data={"value": value})
        except Exception as e:
            return ModuleResult(success=False, error=f"读取公式/值失败: {e}")


@register_executor
class ExcelSetZoomExecutor(ModuleExecutor):
    """设置工作表视图缩放比例（10~400）并可显示/隐藏网格线"""

    @property
    def module_type(self) -> str:
        return "excel_set_zoom"

    async def execute(self, config: dict, context: ExecutionContext) -> ModuleResult:
        path = _resolve_path(context, context.resolve_value(config.get("filePath", "")))
        sheet_name = context.resolve_value(config.get("sheetName", ""))
        zoom = to_int(config.get("zoom", 100), 100, context)
        show_grid = to_bool(config.get("showGridLines", True), context)
        try:
            zoom = max(10, min(400, zoom))
            wb = _load_wb(path)
            ws = _get_ws(wb, sheet_name)
            ws.sheet_view.zoomScale = zoom
            ws.sheet_view.showGridLines = show_grid
            wb.save(path)
            return ModuleResult(success=True, message=f"已设置 {ws.title} 缩放 {zoom}%，网格线{'显示' if show_grid else '隐藏'}")
        except Exception as e:
            return ModuleResult(success=False, error=f"设置视图缩放失败: {e}")
