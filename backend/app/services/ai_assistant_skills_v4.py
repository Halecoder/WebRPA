"""WebRPA 小助手 - v4 超级增强 Skills（通用能力封装）

设计目标：把"AI 高频会用到"的通用 Python 工具能力封装为 Skill，
让 AI 不必每次都生成 python_script 让用户跑——这些是 AI 自己直接调用的工具。

涵盖能力（共 25+ 个 skill）：
  A. 文件 / 文本 / 编码    fast_read_file / fast_write_file / append_file / encode_text
  B. JSON / YAML / TOML   parse_yaml / parse_toml / dump_yaml / pretty_json
  C. HTTP 直接请求         http_get / http_post（不走工作流，AI 自用）
  D. 图片 / 二维码          download_image / qr_encode / qr_decode
  E. 时间 / 日期            now / parse_date / format_date / time_diff
  F. 计算 / 表达式          eval_math / regex_test
  G. 系统信息              get_system_info / get_disk_usage / get_processes_top
  H. Excel/CSV 直读        read_excel_quick / read_csv_quick
  I. 网页静态抓 + 解析      fetch_html / extract_links / extract_text_by_css
  J. AI 实用助手            slugify / hash_text / clipboard_get / clipboard_set
  K. 格式化输出            format_table（把 list[dict] 渲染成漂亮 markdown 表）
"""

from __future__ import annotations

import asyncio
import base64
import hashlib
import json
import os
import platform
import re
import sys
import time
from datetime import datetime, timedelta, timezone
from typing import Any
from urllib.parse import urljoin, urlparse

from app.services.ai_assistant_skills import Skill, registry


# =============================================================================
# A. 文件 / 文本 / 编码
# =============================================================================

async def skill_fast_read_file(
    path: str,
    encoding: str = "utf-8",
    max_chars: int = 80000,
    **_: Any,
) -> dict[str, Any]:
    """快速读文件文本（不走 read_text_file 工作流模块），AI 自用排查/分析。"""
    try:
        if not os.path.exists(path):
            return {"error": f"文件不存在：{path}"}
        with open(path, "r", encoding=encoding, errors="replace") as f:
            text = f.read(max_chars + 1)
        truncated = len(text) > max_chars
        if truncated:
            text = text[:max_chars]
        return {
            "path": os.path.abspath(path),
            "size_bytes": os.path.getsize(path),
            "char_count": len(text),
            "truncated": truncated,
            "content": text,
        }
    except Exception as e:
        return {"error": f"读文件失败：{e}"}


async def skill_fast_write_file(
    path: str,
    content: str,
    encoding: str = "utf-8",
    create_dirs: bool = True,
    **_: Any,
) -> dict[str, Any]:
    """快速写文件文本（覆盖模式）。create_dirs=True 时自动建父目录。"""
    try:
        if create_dirs:
            parent = os.path.dirname(os.path.abspath(path))
            if parent:
                os.makedirs(parent, exist_ok=True)
        with open(path, "w", encoding=encoding) as f:
            f.write(content)
        return {
            "path": os.path.abspath(path),
            "bytes_written": os.path.getsize(path),
            "ok": True,
        }
    except Exception as e:
        return {"error": f"写文件失败：{e}"}


async def skill_append_file(
    path: str,
    content: str,
    encoding: str = "utf-8",
    add_newline: bool = True,
    **_: Any,
) -> dict[str, Any]:
    """追加内容到文件末尾。"""
    try:
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        with open(path, "a", encoding=encoding) as f:
            f.write(content)
            if add_newline and not content.endswith("\n"):
                f.write("\n")
        return {"path": os.path.abspath(path), "ok": True}
    except Exception as e:
        return {"error": f"追加失败：{e}"}


async def skill_encode_text(
    text: str,
    method: str = "base64",
    **_: Any,
) -> dict[str, Any]:
    """文本编解码：method 取 base64 / base64_decode / url / url_decode / hex / hex_decode。"""
    try:
        method = method.lower()
        b = text.encode("utf-8") if isinstance(text, str) else text
        if method == "base64":
            return {"result": base64.b64encode(b).decode("ascii")}
        if method == "base64_decode":
            return {"result": base64.b64decode(text).decode("utf-8", errors="replace")}
        if method in ("url", "urlencode"):
            from urllib.parse import quote
            return {"result": quote(text, safe="")}
        if method in ("url_decode", "urldecode"):
            from urllib.parse import unquote
            return {"result": unquote(text)}
        if method == "hex":
            return {"result": b.hex()}
        if method == "hex_decode":
            return {"result": bytes.fromhex(text).decode("utf-8", errors="replace")}
        return {"error": f"未知 method：{method}"}
    except Exception as e:
        return {"error": f"编解码失败：{e}"}


async def skill_hash_text(
    text: str,
    algorithm: str = "md5",
    **_: Any,
) -> dict[str, Any]:
    """文本哈希：md5 / sha1 / sha256 / sha512。"""
    try:
        algo = algorithm.lower()
        if algo not in ("md5", "sha1", "sha256", "sha512"):
            return {"error": f"不支持的算法：{algorithm}"}
        h = hashlib.new(algo, text.encode("utf-8")).hexdigest()
        return {"algorithm": algo, "hash": h}
    except Exception as e:
        return {"error": f"哈希失败：{e}"}


# =============================================================================
# B. 数据格式 (JSON / YAML / TOML)
# =============================================================================

async def skill_pretty_json(
    text: str,
    indent: int = 2,
    sort_keys: bool = False,
    **_: Any,
) -> dict[str, Any]:
    """JSON 美化（自动检测 / 转化为漂亮缩进字符串）。"""
    try:
        obj = json.loads(text)
        return {"result": json.dumps(obj, ensure_ascii=False, indent=indent, sort_keys=sort_keys)}
    except Exception as e:
        return {"error": f"JSON 解析失败：{e}"}


async def skill_parse_yaml(text: str, **_: Any) -> dict[str, Any]:
    """解析 YAML 字符串为 dict/list。"""
    try:
        try:
            import yaml  # type: ignore
        except ImportError:
            return {"error": "需要 PyYAML：pip install pyyaml"}
        obj = yaml.safe_load(text)
        return {"result": obj}
    except Exception as e:
        return {"error": f"YAML 解析失败：{e}"}


async def skill_dump_yaml(data: Any, **_: Any) -> dict[str, Any]:
    """把 dict/list 转成 YAML 字符串。"""
    try:
        try:
            import yaml  # type: ignore
        except ImportError:
            return {"error": "需要 PyYAML：pip install pyyaml"}
        return {"result": yaml.safe_dump(data, allow_unicode=True, sort_keys=False)}
    except Exception as e:
        return {"error": f"YAML 序列化失败：{e}"}


async def skill_parse_toml(text: str, **_: Any) -> dict[str, Any]:
    """解析 TOML 字符串为 dict（Python 3.11+ 自带 tomllib）。"""
    try:
        try:
            import tomllib  # type: ignore
            obj = tomllib.loads(text)
        except ImportError:
            try:
                import toml  # type: ignore
                obj = toml.loads(text)
            except ImportError:
                return {"error": "需要 toml：pip install toml"}
        return {"result": obj}
    except Exception as e:
        return {"error": f"TOML 解析失败：{e}"}


# =============================================================================
# C. HTTP 直接请求（AI 自用，不走工作流）
# =============================================================================

async def skill_http_get(
    url: str,
    headers: dict | None = None,
    params: dict | None = None,
    timeout: int = 15,
    max_chars: int = 30000,
    **_: Any,
) -> dict[str, Any]:
    """AI 直接发 HTTP GET（用于查 API、抓静态页等）。响应文本最多保留 30000 字符。"""
    try:
        try:
            import httpx  # type: ignore
        except ImportError:
            return {"error": "需要 httpx：pip install httpx"}
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            r = await client.get(url, headers=headers, params=params)
            text = r.text
            truncated = len(text) > max_chars
            if truncated:
                text = text[:max_chars]
            return {
                "status_code": r.status_code,
                "headers": dict(r.headers),
                "url": str(r.url),
                "text": text,
                "truncated": truncated,
                "content_type": r.headers.get("content-type", ""),
            }
    except Exception as e:
        return {"error": f"GET 失败：{e}"}


async def skill_http_post(
    url: str,
    data: Any | None = None,
    json_body: Any | None = None,
    headers: dict | None = None,
    timeout: int = 30,
    max_chars: int = 30000,
    **_: Any,
) -> dict[str, Any]:
    """AI 直接发 HTTP POST。data 用表单，json_body 用 JSON 体。"""
    try:
        try:
            import httpx  # type: ignore
        except ImportError:
            return {"error": "需要 httpx：pip install httpx"}
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            r = await client.post(url, data=data, json=json_body, headers=headers)
            text = r.text
            truncated = len(text) > max_chars
            if truncated:
                text = text[:max_chars]
            return {
                "status_code": r.status_code,
                "headers": dict(r.headers),
                "text": text,
                "truncated": truncated,
            }
    except Exception as e:
        return {"error": f"POST 失败：{e}"}


# =============================================================================
# D. 图片 / 二维码
# =============================================================================

async def skill_download_image(
    url: str,
    save_path: str | None = None,
    timeout: int = 20,
    **_: Any,
) -> dict[str, Any]:
    """下载图片到本地（不存就返回 base64）。"""
    try:
        try:
            import httpx  # type: ignore
        except ImportError:
            return {"error": "需要 httpx：pip install httpx"}
        async with httpx.AsyncClient(timeout=timeout, follow_redirects=True) as client:
            r = await client.get(url)
            if r.status_code != 200:
                return {"error": f"HTTP {r.status_code}"}
            img_bytes = r.content
            if save_path:
                os.makedirs(os.path.dirname(os.path.abspath(save_path)), exist_ok=True)
                with open(save_path, "wb") as f:
                    f.write(img_bytes)
                return {
                    "saved_to": os.path.abspath(save_path),
                    "size_bytes": len(img_bytes),
                    "content_type": r.headers.get("content-type", ""),
                }
            return {
                "base64": base64.b64encode(img_bytes).decode("ascii"),
                "size_bytes": len(img_bytes),
                "content_type": r.headers.get("content-type", ""),
            }
    except Exception as e:
        return {"error": f"下载失败：{e}"}


async def skill_qr_encode(
    text: str,
    save_path: str | None = None,
    box_size: int = 10,
    **_: Any,
) -> dict[str, Any]:
    """生成二维码：默认返回 base64 PNG，传 save_path 则保存。"""
    try:
        try:
            import qrcode  # type: ignore
            from io import BytesIO
        except ImportError:
            return {"error": "需要 qrcode：pip install qrcode[pil]"}
        qr = qrcode.QRCode(box_size=box_size, border=2)
        qr.add_data(text)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        if save_path:
            os.makedirs(os.path.dirname(os.path.abspath(save_path)), exist_ok=True)
            img.save(save_path)
            return {"saved_to": os.path.abspath(save_path)}
        buf = BytesIO()
        img.save(buf, format="PNG")
        return {"base64_png": base64.b64encode(buf.getvalue()).decode("ascii")}
    except Exception as e:
        return {"error": f"生成失败：{e}"}


async def skill_qr_decode(
    image_path: str,
    **_: Any,
) -> dict[str, Any]:
    """解码二维码图片。"""
    try:
        if not os.path.exists(image_path):
            return {"error": "图片不存在"}
        try:
            from PIL import Image  # type: ignore
            from pyzbar.pyzbar import decode  # type: ignore
        except ImportError:
            return {"error": "需要：pip install pyzbar pillow"}
        img = Image.open(image_path)
        results = decode(img)
        if not results:
            return {"error": "未识别到二维码"}
        return {
            "results": [
                {"type": r.type, "data": r.data.decode("utf-8", errors="replace")}
                for r in results
            ]
        }
    except Exception as e:
        return {"error": f"解码失败：{e}"}


# =============================================================================
# E. 时间 / 日期
# =============================================================================

async def skill_now(timezone_offset: int | None = None, **_: Any) -> dict[str, Any]:
    """当前时间。timezone_offset 是相对 UTC 的小时偏移，默认本地时区。"""
    if timezone_offset is None:
        now = datetime.now()
    else:
        now = datetime.now(timezone(timedelta(hours=timezone_offset)))
    return {
        "iso": now.isoformat(),
        "timestamp_ms": int(now.timestamp() * 1000),
        "timestamp_s": int(now.timestamp()),
        "year": now.year, "month": now.month, "day": now.day,
        "hour": now.hour, "minute": now.minute, "second": now.second,
        "weekday": now.weekday(),  # 0=Monday
        "weekday_zh": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][now.weekday()],
    }


async def skill_format_date(
    iso: str = "",
    timestamp: int | None = None,
    fmt: str = "%Y-%m-%d %H:%M:%S",
    **_: Any,
) -> dict[str, Any]:
    """把 ISO 字符串或 Unix 时间戳格式化成想要的字符串。"""
    try:
        if timestamp is not None:
            dt = datetime.fromtimestamp(int(timestamp))
        elif iso:
            dt = datetime.fromisoformat(iso)
        else:
            return {"error": "需要 iso 或 timestamp 之一"}
        return {"result": dt.strftime(fmt)}
    except Exception as e:
        return {"error": f"格式化失败：{e}"}


async def skill_time_diff(
    start: str,
    end: str = "",
    **_: Any,
) -> dict[str, Any]:
    """计算两个 ISO 时间字符串的差值（end 默认 now）。"""
    try:
        s = datetime.fromisoformat(start)
        e = datetime.fromisoformat(end) if end else datetime.now()
        delta = e - s
        seconds = int(delta.total_seconds())
        return {
            "seconds": seconds,
            "minutes": round(seconds / 60, 2),
            "hours": round(seconds / 3600, 2),
            "days": delta.days,
            "human": _human_duration(seconds),
        }
    except Exception as e:
        return {"error": f"时间差计算失败：{e}"}


def _human_duration(seconds: int) -> str:
    if seconds < 60:
        return f"{seconds} 秒"
    if seconds < 3600:
        return f"{seconds // 60} 分 {seconds % 60} 秒"
    if seconds < 86400:
        h = seconds // 3600
        m = (seconds % 3600) // 60
        return f"{h} 时 {m} 分"
    d = seconds // 86400
    h = (seconds % 86400) // 3600
    return f"{d} 天 {h} 时"


# =============================================================================
# F. 计算 / 正则
# =============================================================================

async def skill_eval_math(expression: str, **_: Any) -> dict[str, Any]:
    """安全计算数学表达式（+ - * / ** % () 以及 math.* 函数）。"""
    try:
        import math
        allowed = {k: getattr(math, k) for k in dir(math) if not k.startswith("_")}
        allowed["abs"] = abs
        allowed["min"] = min
        allowed["max"] = max
        allowed["round"] = round
        allowed["sum"] = sum
        # 严禁导入和文件操作
        if any(k in expression for k in ("import", "__", "open", "exec", "eval", "os.", "sys.")):
            return {"error": "表达式包含禁用关键字"}
        result = eval(expression, {"__builtins__": {}}, allowed)
        return {"result": result, "expression": expression}
    except Exception as e:
        return {"error": f"计算失败：{e}"}


async def skill_regex_test(
    pattern: str,
    text: str,
    flags: str = "",
    **_: Any,
) -> dict[str, Any]:
    """正则测试：返回所有匹配 + 第一个 group。flags 支持 i/m/s/x。"""
    try:
        re_flags = 0
        for ch in flags.lower():
            if ch == "i": re_flags |= re.IGNORECASE
            elif ch == "m": re_flags |= re.MULTILINE
            elif ch == "s": re_flags |= re.DOTALL
            elif ch == "x": re_flags |= re.VERBOSE
        matches = list(re.finditer(pattern, text, re_flags))
        return {
            "match_count": len(matches),
            "matches": [
                {
                    "match": m.group(),
                    "start": m.start(),
                    "end": m.end(),
                    "groups": m.groups(),
                    "groupdict": m.groupdict(),
                }
                for m in matches[:50]
            ],
        }
    except Exception as e:
        return {"error": f"正则错误：{e}"}


# =============================================================================
# G. 系统信息
# =============================================================================

async def skill_get_system_info(**_: Any) -> dict[str, Any]:
    """系统基础信息（OS、CPU、内存）。"""
    try:
        info = {
            "os": platform.system(),
            "os_version": platform.version(),
            "platform": platform.platform(),
            "python_version": sys.version,
            "cpu_count": os.cpu_count(),
        }
        try:
            import psutil  # type: ignore
            mem = psutil.virtual_memory()
            info.update({
                "cpu_percent": psutil.cpu_percent(interval=0.3),
                "memory_total_gb": round(mem.total / 1024**3, 2),
                "memory_available_gb": round(mem.available / 1024**3, 2),
                "memory_percent": mem.percent,
            })
        except ImportError:
            pass
        return info
    except Exception as e:
        return {"error": f"获取失败：{e}"}


async def skill_get_disk_usage(path: str = "C:\\", **_: Any) -> dict[str, Any]:
    """获取指定路径的磁盘使用情况。"""
    try:
        import shutil
        total, used, free = shutil.disk_usage(path)
        return {
            "path": path,
            "total_gb": round(total / 1024**3, 2),
            "used_gb": round(used / 1024**3, 2),
            "free_gb": round(free / 1024**3, 2),
            "percent_used": round(used * 100 / total, 1),
        }
    except Exception as e:
        return {"error": f"获取失败：{e}"}


async def skill_get_processes_top(
    top: int = 10,
    sort_by: str = "memory",
    **_: Any,
) -> dict[str, Any]:
    """列出 CPU 或内存占用前 N 的进程。sort_by: cpu / memory。"""
    try:
        try:
            import psutil  # type: ignore
        except ImportError:
            return {"error": "需要 psutil：pip install psutil"}
        procs = []
        for p in psutil.process_iter(['pid', 'name', 'memory_percent', 'cpu_percent']):
            try:
                procs.append(p.info)
            except Exception:
                continue
        key = 'memory_percent' if sort_by == "memory" else 'cpu_percent'
        procs.sort(key=lambda x: x.get(key) or 0, reverse=True)
        return {"processes": procs[: int(top)]}
    except Exception as e:
        return {"error": f"获取失败：{e}"}


# =============================================================================
# H. Excel / CSV 直读（AI 排查 / 给用户预览用，不入工作流）
# =============================================================================

async def skill_read_excel_quick(
    path: str,
    sheet: str | int = 0,
    max_rows: int = 50,
    **_: Any,
) -> dict[str, Any]:
    """快读 Excel 前 N 行（AI 用来快速了解表结构）。"""
    try:
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            return {"error": "需要 openpyxl：pip install openpyxl"}
        if not os.path.exists(path):
            return {"error": "文件不存在"}
        wb = load_workbook(path, read_only=True, data_only=True)
        if isinstance(sheet, int):
            ws = wb.worksheets[sheet]
        else:
            ws = wb[sheet]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))
        return {
            "sheet_name": ws.title,
            "all_sheets": wb.sheetnames,
            "row_count_loaded": len(rows),
            "rows": rows,
        }
    except Exception as e:
        return {"error": f"读 Excel 失败：{e}"}


async def skill_read_csv_quick(
    path: str,
    encoding: str = "utf-8-sig",
    max_rows: int = 100,
    delimiter: str = ",",
    **_: Any,
) -> dict[str, Any]:
    """快读 CSV 前 N 行（自动识别 BOM）。"""
    try:
        import csv
        if not os.path.exists(path):
            return {"error": "文件不存在"}
        with open(path, "r", encoding=encoding, errors="replace", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            rows = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                rows.append(row)
        return {"row_count_loaded": len(rows), "rows": rows}
    except Exception as e:
        return {"error": f"读 CSV 失败：{e}"}


# =============================================================================
# I. 网页抓 + 解析
# =============================================================================

async def skill_extract_links(
    url: str = "",
    html: str = "",
    same_domain_only: bool = True,
    **_: Any,
) -> dict[str, Any]:
    """从 URL 或 HTML 中提取所有链接。"""
    try:
        try:
            from bs4 import BeautifulSoup  # type: ignore
        except ImportError:
            return {"error": "需要 beautifulsoup4：pip install beautifulsoup4"}
        if url and not html:
            try:
                import httpx  # type: ignore
            except ImportError:
                return {"error": "需要 httpx"}
            async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
                r = await client.get(url)
                html = r.text
                base_url = str(r.url)
        else:
            base_url = url
        if not html:
            return {"error": "html 和 url 至少给一个"}
        soup = BeautifulSoup(html, "html.parser")
        links = []
        seen = set()
        domain = urlparse(base_url).netloc if base_url else ""
        for a in soup.find_all("a", href=True):
            href = a["href"]
            absolute = urljoin(base_url, href) if base_url else href
            if same_domain_only and domain and urlparse(absolute).netloc and urlparse(absolute).netloc != domain:
                continue
            if absolute in seen:
                continue
            seen.add(absolute)
            links.append({"text": a.get_text(strip=True)[:120], "url": absolute})
            if len(links) >= 200:
                break
        return {"link_count": len(links), "links": links}
    except Exception as e:
        return {"error": f"提取失败：{e}"}


async def skill_extract_text_by_css(
    url: str = "",
    html: str = "",
    selector: str = "",
    **_: Any,
) -> dict[str, Any]:
    """从 URL 或 HTML 中按 CSS selector 抓出所有匹配元素的纯文本。"""
    try:
        try:
            from bs4 import BeautifulSoup  # type: ignore
        except ImportError:
            return {"error": "需要 beautifulsoup4"}
        if not selector:
            return {"error": "selector 不能为空"}
        if url and not html:
            try:
                import httpx  # type: ignore
            except ImportError:
                return {"error": "需要 httpx"}
            async with httpx.AsyncClient(timeout=15, follow_redirects=True) as client:
                r = await client.get(url)
                html = r.text
        if not html:
            return {"error": "html 和 url 至少给一个"}
        soup = BeautifulSoup(html, "html.parser")
        elems = soup.select(selector)
        texts = [e.get_text(strip=True) for e in elems if e.get_text(strip=True)]
        return {"count": len(texts), "texts": texts[:200]}
    except Exception as e:
        return {"error": f"提取失败：{e}"}


# =============================================================================
# J. 实用工具
# =============================================================================

async def skill_slugify(
    text: str,
    separator: str = "-",
    lowercase: bool = True,
    **_: Any,
) -> dict[str, Any]:
    """文件名/URL 友好化（去掉特殊字符 + 中文音译）。"""
    try:
        try:
            from slugify import slugify as _slug  # type: ignore
            return {"result": _slug(text, separator=separator, lowercase=lowercase)}
        except ImportError:
            # 无 python-slugify 时降级处理
            t = re.sub(r"[\s\u3000]+", separator, text.strip())
            t = re.sub(rf"[^\w\u4e00-\u9fa5{re.escape(separator)}]", "", t)
            return {"result": t.lower() if lowercase else t, "note": "未装 python-slugify，降级处理"}
    except Exception as e:
        return {"error": f"slugify 失败：{e}"}


async def skill_clipboard_get(**_: Any) -> dict[str, Any]:
    """读取系统剪贴板内容（AI 自用）。"""
    try:
        try:
            import pyperclip  # type: ignore
        except ImportError:
            return {"error": "需要 pyperclip"}
        return {"text": pyperclip.paste()}
    except Exception as e:
        return {"error": f"读取剪贴板失败：{e}"}


async def skill_clipboard_set(text: str, **_: Any) -> dict[str, Any]:
    """写入系统剪贴板。"""
    try:
        try:
            import pyperclip  # type: ignore
        except ImportError:
            return {"error": "需要 pyperclip"}
        pyperclip.copy(text)
        return {"ok": True, "char_count": len(text)}
    except Exception as e:
        return {"error": f"写剪贴板失败：{e}"}


async def skill_sleep(
    seconds: float = 1.0,
    reason: str = "",
    **_: Any,
) -> dict[str, Any]:
    """AI 自己等待 N 秒（真正的阻塞等待）。

    sleep 完成后才会返回，AI 在 N 秒后才会拿到结果继续下一步。
    最常用场景：run_workflow 后等几秒再 get_logs 看结果。

    上限 300 秒。需要更长等待请用 schedule_one_shot 延时任务。
    """
    try:
        s = float(seconds)
        if s < 0:
            s = 0
        if s > 300:
            return {
                "error": f"sleep 上限 300 秒，传入 {s} 秒过大。"
                         "更长等待请用 schedule_one_shot 延时任务。"
            }
        await asyncio.sleep(s)
        return {
            "slept_seconds": s,
            "reason": reason or "ok",
            "note": "已真实等待完成，可继续下一步操作",
        }
    except Exception as e:
        return {"error": f"sleep 失败：{e}"}


async def skill_format_table(
    rows: list[dict[str, Any]],
    columns: list[str] | None = None,
    max_col_width: int = 50,
    **_: Any,
) -> dict[str, Any]:
    """把 list[dict] 渲染成 markdown 表格字符串（AI 总结报告时用）。"""
    try:
        if not rows:
            return {"result": "_(空表)_"}
        cols = columns or list(rows[0].keys())
        def _fmt(v: Any) -> str:
            s = "" if v is None else str(v)
            s = s.replace("|", "\\|").replace("\n", " ")
            return s[:max_col_width]
        lines = []
        lines.append("| " + " | ".join(cols) + " |")
        lines.append("| " + " | ".join(["---"] * len(cols)) + " |")
        for r in rows:
            lines.append("| " + " | ".join(_fmt(r.get(c)) for c in cols) + " |")
        return {"result": "\n".join(lines)}
    except Exception as e:
        return {"error": f"格式化失败：{e}"}


async def skill_run_python_inline(
    code: str,
    timeout: int = 30,
    **_: Any,
) -> dict[str, Any]:
    """直接运行一段 Python 代码（AI 自用快速计算/处理，不走 python_script 模块）。

    用 subprocess 隔离运行（独立 Python313 环境，最长 30 秒）。
    适合：临时计算、字符串处理、调用 stdlib、做 ad-hoc 分析。
    不适合：需要工作流变量、需要返回到画布的复杂任务。
    """
    try:
        import subprocess
        import tempfile
        # 写到临时文件
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".py", delete=False, encoding="utf-8"
        ) as f:
            f.write(code)
            tmp = f.name
        # 找内置 Python313
        from pathlib import Path
        root = Path(__file__).parent.parent.parent.parent
        py = root / "Python313" / "python.exe"
        py_path = str(py) if py.exists() else sys.executable
        try:
            proc = await asyncio.create_subprocess_exec(
                py_path, tmp,
                stdout=subprocess.PIPE,
                stderr=subprocess.PIPE,
            )
            try:
                out, err = await asyncio.wait_for(proc.communicate(), timeout=int(timeout))
            except asyncio.TimeoutError:
                try: proc.kill()
                except Exception: pass
                return {"error": f"超时 {timeout}s"}
            return {
                "return_code": proc.returncode,
                "stdout": out.decode("utf-8", errors="replace"),
                "stderr": err.decode("utf-8", errors="replace"),
            }
        finally:
            try: os.unlink(tmp)
            except Exception: pass
    except Exception as e:
        return {"error": f"执行失败：{e}"}


# =============================================================================
# 注册全部 skill
# =============================================================================

def _register_v4() -> None:
    # A. 文件 / 文本
    registry.register(Skill(
        name="fast_read_file",
        description="快速读取本地文件文本（最多 80000 字符），AI 自用排查。比 client_action(read_text_file) 快很多。",
        parameters={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "encoding": {"type": "string", "default": "utf-8"},
                "max_chars": {"type": "integer", "default": 80000},
            },
            "required": ["path"],
        },
        handler=skill_fast_read_file,
    ))
    registry.register(Skill(
        name="fast_write_file",
        description="快速写本地文件文本（覆盖模式，自动建父目录）。",
        parameters={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "content": {"type": "string"},
                "encoding": {"type": "string", "default": "utf-8"},
            },
            "required": ["path", "content"],
        },
        handler=skill_fast_write_file,
        requires_approval=True,
    ))
    registry.register(Skill(
        name="append_file",
        description="追加内容到文件末尾。",
        parameters={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "content": {"type": "string"},
            },
            "required": ["path", "content"],
        },
        handler=skill_append_file,
        requires_approval=True,
    ))
    registry.register(Skill(
        name="encode_text",
        description="文本编解码：base64 / base64_decode / url / url_decode / hex / hex_decode。",
        parameters={
            "type": "object",
            "properties": {
                "text": {"type": "string"},
                "method": {"type": "string"},
            },
            "required": ["text", "method"],
        },
        handler=skill_encode_text,
    ))
    registry.register(Skill(
        name="hash_text",
        description="文本哈希：md5 / sha1 / sha256 / sha512。",
        parameters={
            "type": "object",
            "properties": {
                "text": {"type": "string"},
                "algorithm": {"type": "string", "default": "md5"},
            },
            "required": ["text"],
        },
        handler=skill_hash_text,
    ))

    # B. 数据格式
    registry.register(Skill(
        name="pretty_json",
        description="JSON 美化（输入字符串，输出漂亮缩进版）。",
        parameters={"type": "object", "properties": {"text": {"type": "string"}}, "required": ["text"]},
        handler=skill_pretty_json,
    ))
    registry.register(Skill(
        name="parse_yaml",
        description="解析 YAML 字符串为 dict/list。",
        parameters={"type": "object", "properties": {"text": {"type": "string"}}, "required": ["text"]},
        handler=skill_parse_yaml,
    ))
    registry.register(Skill(
        name="dump_yaml",
        description="把 dict/list 转成 YAML 字符串。",
        parameters={"type": "object", "properties": {"data": {}}, "required": ["data"]},
        handler=skill_dump_yaml,
    ))
    registry.register(Skill(
        name="parse_toml",
        description="解析 TOML 字符串为 dict。",
        parameters={"type": "object", "properties": {"text": {"type": "string"}}, "required": ["text"]},
        handler=skill_parse_toml,
    ))

    # C. HTTP
    registry.register(Skill(
        name="http_get",
        description="AI 直接发 HTTP GET（查 API / 抓静态页）。响应文本最多 30000 字符。",
        parameters={
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "headers": {"type": "object"},
                "params": {"type": "object"},
                "timeout": {"type": "integer", "default": 15},
            },
            "required": ["url"],
        },
        handler=skill_http_get,
    ))
    registry.register(Skill(
        name="http_post",
        description="AI 直接发 HTTP POST（json_body 走 JSON / data 走 form）。",
        parameters={
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "data": {},
                "json_body": {},
                "headers": {"type": "object"},
                "timeout": {"type": "integer", "default": 30},
            },
            "required": ["url"],
        },
        handler=skill_http_post,
    ))

    # D. 图片 / 二维码
    registry.register(Skill(
        name="download_image",
        description="下载图片到本地（不传 save_path 则返回 base64）。",
        parameters={
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "save_path": {"type": "string"},
            },
            "required": ["url"],
        },
        handler=skill_download_image,
    ))
    registry.register(Skill(
        name="qr_encode",
        description="生成二维码（默认返回 base64 PNG，传 save_path 则保存到本地）。",
        parameters={
            "type": "object",
            "properties": {
                "text": {"type": "string"},
                "save_path": {"type": "string"},
                "box_size": {"type": "integer", "default": 10},
            },
            "required": ["text"],
        },
        handler=skill_qr_encode,
    ))
    registry.register(Skill(
        name="qr_decode",
        description="解码二维码图片。",
        parameters={
            "type": "object",
            "properties": {"image_path": {"type": "string"}},
            "required": ["image_path"],
        },
        handler=skill_qr_decode,
    ))

    # E. 时间
    registry.register(Skill(
        name="now",
        description="获取当前时间（ISO/timestamp/星期）。",
        parameters={
            "type": "object",
            "properties": {"timezone_offset": {"type": "integer"}},
        },
        handler=skill_now,
    ))
    registry.register(Skill(
        name="format_date",
        description="格式化日期：传 iso 字符串或 timestamp，输出指定格式。",
        parameters={
            "type": "object",
            "properties": {
                "iso": {"type": "string"},
                "timestamp": {"type": "integer"},
                "fmt": {"type": "string", "default": "%Y-%m-%d %H:%M:%S"},
            },
        },
        handler=skill_format_date,
    ))
    registry.register(Skill(
        name="time_diff",
        description="计算两个 ISO 时间的差值（end 默认当前时间）。",
        parameters={
            "type": "object",
            "properties": {
                "start": {"type": "string"},
                "end": {"type": "string"},
            },
            "required": ["start"],
        },
        handler=skill_time_diff,
    ))

    # F. 计算 / 正则
    registry.register(Skill(
        name="eval_math",
        description="安全计算数学表达式（支持 math.* 函数，禁用文件/导入操作）。",
        parameters={
            "type": "object",
            "properties": {"expression": {"type": "string"}},
            "required": ["expression"],
        },
        handler=skill_eval_math,
    ))
    registry.register(Skill(
        name="regex_test",
        description="正则测试：返回所有匹配 + 分组。flags 支持 i/m/s/x。",
        parameters={
            "type": "object",
            "properties": {
                "pattern": {"type": "string"},
                "text": {"type": "string"},
                "flags": {"type": "string", "default": ""},
            },
            "required": ["pattern", "text"],
        },
        handler=skill_regex_test,
    ))

    # G. 系统信息
    registry.register(Skill(
        name="get_system_info",
        description="系统基础信息（OS/CPU/内存）。",
        parameters={"type": "object", "properties": {}},
        handler=skill_get_system_info,
    ))
    registry.register(Skill(
        name="get_disk_usage",
        description="磁盘使用情况（默认 C: 盘）。",
        parameters={
            "type": "object",
            "properties": {"path": {"type": "string", "default": "C:\\"}},
        },
        handler=skill_get_disk_usage,
    ))
    registry.register(Skill(
        name="get_processes_top",
        description="按 CPU 或内存占用列前 N 个进程。",
        parameters={
            "type": "object",
            "properties": {
                "top": {"type": "integer", "default": 10},
                "sort_by": {"type": "string", "default": "memory"},
            },
        },
        handler=skill_get_processes_top,
    ))

    # H. Excel/CSV 直读
    registry.register(Skill(
        name="read_excel_quick",
        description="快读 Excel 前 N 行（AI 用来探查表结构）。",
        parameters={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "sheet": {},
                "max_rows": {"type": "integer", "default": 50},
            },
            "required": ["path"],
        },
        handler=skill_read_excel_quick,
    ))
    registry.register(Skill(
        name="read_csv_quick",
        description="快读 CSV 前 N 行。",
        parameters={
            "type": "object",
            "properties": {
                "path": {"type": "string"},
                "encoding": {"type": "string", "default": "utf-8-sig"},
                "max_rows": {"type": "integer", "default": 100},
                "delimiter": {"type": "string", "default": ","},
            },
            "required": ["path"],
        },
        handler=skill_read_csv_quick,
    ))

    # I. 网页解析
    registry.register(Skill(
        name="extract_links",
        description="从 URL 或 HTML 中提取所有链接（默认只取同域）。",
        parameters={
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "html": {"type": "string"},
                "same_domain_only": {"type": "boolean", "default": True},
            },
        },
        handler=skill_extract_links,
    ))
    registry.register(Skill(
        name="extract_text_by_css",
        description="从 URL 或 HTML 中按 CSS selector 抓出元素纯文本。",
        parameters={
            "type": "object",
            "properties": {
                "url": {"type": "string"},
                "html": {"type": "string"},
                "selector": {"type": "string"},
            },
            "required": ["selector"],
        },
        handler=skill_extract_text_by_css,
    ))

    # J. 实用工具
    registry.register(Skill(
        name="slugify",
        description="文件名/URL 友好化（去特殊字符 + 中文音译）。",
        parameters={
            "type": "object",
            "properties": {
                "text": {"type": "string"},
                "separator": {"type": "string", "default": "-"},
                "lowercase": {"type": "boolean", "default": True},
            },
            "required": ["text"],
        },
        handler=skill_slugify,
    ))
    registry.register(Skill(
        name="clipboard_get",
        description="读取系统剪贴板（AI 自用，不走工作流）。",
        parameters={"type": "object", "properties": {}},
        handler=skill_clipboard_get,
    ))
    registry.register(Skill(
        name="clipboard_set",
        description="写入系统剪贴板。",
        parameters={
            "type": "object",
            "properties": {"text": {"type": "string"}},
            "required": ["text"],
        },
        handler=skill_clipboard_set,
    ))
    registry.register(Skill(
        name="sleep",
        description=(
            "AI 自己阻塞等待 N 秒后返回（真实等待，N 秒后才会执行下一步工具）。"
            "最常用场景：run_workflow 后等几秒再 get_logs。上限 300 秒。"
            "示例：sleep(seconds=3, reason='等工作流跑完')"
        ),
        parameters={
            "type": "object",
            "properties": {
                "seconds": {"type": "number", "description": "等待秒数（0~300）"},
                "reason": {"type": "string", "description": "等待原因（仅日志，可选）"},
            },
            "required": ["seconds"],
        },
        handler=skill_sleep,
    ))
    registry.register(Skill(
        name="format_table",
        description="把 list[dict] 渲染成 markdown 表格字符串（总结报告用）。",
        parameters={
            "type": "object",
            "properties": {
                "rows": {"type": "array"},
                "columns": {"type": "array"},
                "max_col_width": {"type": "integer", "default": 50},
            },
            "required": ["rows"],
        },
        handler=skill_format_table,
    ))
    registry.register(Skill(
        name="run_python_inline",
        description="直接执行一段 Python 代码（AI 自用快速计算，最多 30 秒）。不走画布工作流。",
        parameters={
            "type": "object",
            "properties": {
                "code": {"type": "string"},
                "timeout": {"type": "integer", "default": 30},
            },
            "required": ["code"],
        },
        handler=skill_run_python_inline,
        requires_approval=True,
    ))


_register_v4()
