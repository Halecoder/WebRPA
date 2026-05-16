"""数据收集器 - 使用Polars管理和导出数据"""
import os
from pathlib import Path
from typing import Any, Optional
from datetime import datetime

import polars as pl


class DataCollector:
    """数据收集器"""
    
    def __init__(self):
        self.data: list[dict[str, Any]] = []
        self.columns: list[str] = []
        self._current_row: dict[str, Any] = {}
    
    def add_value(self, column: str, value: Any):
        """添加单个值到当前行"""
        if column not in self.columns:
            self.columns.append(column)
        self._current_row[column] = value
    
    def add_row(self, row: dict[str, Any]):
        """添加一行数据"""
        for key in row:
            if key not in self.columns:
                self.columns.append(key)
        self.data.append(row)
    
    def commit_row(self):
        """提交当前行"""
        if self._current_row:
            self.add_row(self._current_row.copy())
            self._current_row = {}
    
    def clear(self):
        """清空数据"""
        self.data = []
        self.columns = []
        self._current_row = {}
    
    def to_dataframe(self) -> pl.DataFrame:
        """转换为Polars DataFrame
        
        把不能直接写入 Excel 的复杂类型（dict/list/tuple/set）序列化为 JSON 字符串，
        保留 datetime/date/time 等原生类型，避免 Excel 失去日期格式过滤能力。
        """
        import json
        from datetime import datetime, date, time
        if not self.data:
            return pl.DataFrame()
        
        def _normalize(v):
            if v is None:
                return None
            if isinstance(v, (str, int, float, bool)):
                return v
            # 保留 datetime/date/time 等原生类型，让 polars/Excel 自己识别
            if isinstance(v, (datetime, date, time)):
                return v
            if isinstance(v, (list, dict, tuple, set)):
                try:
                    return json.dumps(v, ensure_ascii=False, default=str)
                except Exception:
                    return str(v)
            # 其他对象（Path 等）转字符串
            return str(v)
        
        # 确保所有行都有相同的列
        normalized_data = []
        for row in self.data:
            normalized_row = {col: _normalize(row.get(col)) for col in self.columns}
            normalized_data.append(normalized_row)
        
        return pl.DataFrame(normalized_data)
    
    def to_excel(self, filepath: str, sheet_name: str = '数据') -> str:
        """导出为Excel文件（带样式）
        
        Args:
            filepath: 文件路径
            sheet_name: Sheet名称，默认为'数据'
        """
        df = self.to_dataframe()
        
        # 确保目录存在
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        
        # 导出Excel（带样式）
        _write_styled_excel(df, filepath, sheet_name)
        
        # 写入后立即校验文件存在且非空
        p = Path(filepath)
        if not p.exists():
            raise IOError(f"导出后文件不存在: {filepath}")
        if p.stat().st_size == 0:
            raise IOError(f"导出文件大小为 0: {filepath}")
        
        return filepath
    
    def to_csv(self, filepath: str) -> str:
        """导出为CSV文件"""
        df = self.to_dataframe()
        
        # 确保目录存在
        Path(filepath).parent.mkdir(parents=True, exist_ok=True)
        
        # 导出CSV
        df.write_csv(filepath)
        
        return filepath
    
    @property
    def row_count(self) -> int:
        """获取行数"""
        return len(self.data)
    
    @property
    def column_count(self) -> int:
        """获取列数"""
        return len(self.columns)


def _write_styled_excel(df: pl.DataFrame, filepath: str, sheet_name: str = '数据'):
    """写入带样式的Excel文件
    
    Args:
        df: Polars DataFrame
        filepath: 文件路径
        sheet_name: Sheet名称，默认为'数据'
    """
    # 大数据量阈值：超过此值时使用快速路径（仅头部样式 + 极简单元格写入）
    LARGE_DATA_THRESHOLD = 500
    
    try:
        from xlsxwriter import Workbook
        import openpyxl
        
        # 检查文件是否已存在
        file_exists = Path(filepath).exists()
        row_count = len(df)
        is_large = row_count >= LARGE_DATA_THRESHOLD
        
        if file_exists:
            # 文件存在，使用openpyxl追加或更新Sheet
            try:
                wb = openpyxl.load_workbook(filepath)
                
                # 如果Sheet已存在，删除它
                if sheet_name in wb.sheetnames:
                    del wb[sheet_name]
                
                # 创建新的Sheet
                ws = wb.create_sheet(sheet_name)
                
                # 写入数据（直接使用 polars 转换为列表，避免引入 pandas）
                columns = df.columns
                
                # 写入表头
                header_font = openpyxl.styles.Font(bold=True, color="FFFFFF", size=11)
                header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_align = openpyxl.styles.Alignment(horizontal="center", vertical="center")
                header_border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin', color='2F5496'),
                    right=openpyxl.styles.Side(style='thin', color='2F5496'),
                    top=openpyxl.styles.Side(style='thin', color='2F5496'),
                    bottom=openpyxl.styles.Side(style='thin', color='2F5496')
                )
                for col_idx, col_name in enumerate(columns, 1):
                    cell = ws.cell(row=1, column=col_idx, value=col_name)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_align
                    cell.border = header_border
                
                # 写入数据行
                if is_large:
                    # 大数据量快速路径：批量 append 不写单元格样式
                    for row_data in df.iter_rows():
                        ws.append(list(row_data))
                else:
                    even_fill = openpyxl.styles.PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
                    cell_border = openpyxl.styles.Border(
                        left=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                        right=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                        top=openpyxl.styles.Side(style='thin', color='D9D9D9'),
                        bottom=openpyxl.styles.Side(style='thin', color='D9D9D9')
                    )
                    for row_idx, row_data in enumerate(df.iter_rows(), 2):
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            if row_idx % 2 == 0:
                                cell.fill = even_fill
                            cell.border = cell_border
                
                # 自动调整列宽（仅采样前 200 行计算列宽，避免大数据量遍历）
                sample_rows = list(df.head(min(200, row_count)).iter_rows())
                for col_idx, col_name in enumerate(columns, 1):
                    max_len = len(str(col_name))
                    for r in sample_rows:
                        v = r[col_idx - 1]
                        l = len(str(v)) if v is not None else 0
                        if l > max_len:
                            max_len = l
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_len + 4, 50)
                
                # 冻结首行
                ws.freeze_panes = 'A2'
                
                # 保存工作簿
                wb.save(filepath)
                return
                
            except PermissionError as pe:
                # 文件被 Excel 等程序占用，明确抛出
                raise PermissionError(f"文件被占用，无法写入（请关闭已打开的 Excel 文件）: {filepath}") from pe
            except Exception as e:
                # 如果使用openpyxl失败，回退到xlsxwriter（会覆盖整个文件）
                print(f"[警告] 使用openpyxl追加Sheet失败: {e}，将使用xlsxwriter创建新文件")
                file_exists = False
        
        if not file_exists:
            # 文件不存在，使用xlsxwriter创建新文件
            workbook = Workbook(filepath)
            try:
                worksheet = workbook.add_worksheet(sheet_name)
                
                # 定义样式
                header_format = workbook.add_format({
                    'bold': True,
                    'font_size': 11,
                    'font_color': 'white',
                    'bg_color': '#4472C4',
                    'border': 1,
                    'border_color': '#2F5496',
                    'align': 'center',
                    'valign': 'vcenter',
                    'text_wrap': True,
                })
                
                # 写入表头
                columns = df.columns
                for col_idx, col_name in enumerate(columns):
                    worksheet.write(0, col_idx, col_name, header_format)
                
                if is_large:
                    # 大数据量快速路径：不带逐单元格样式
                    for row_idx, row in enumerate(df.iter_rows()):
                        for col_idx, value in enumerate(row):
                            worksheet.write(row_idx + 1, col_idx, value)
                else:
                    cell_format = workbook.add_format({
                        'font_size': 10,
                        'border': 1,
                        'border_color': '#D9D9D9',
                        'align': 'left',
                        'valign': 'vcenter',
                    })
                    alt_cell_format = workbook.add_format({
                        'font_size': 10,
                        'border': 1,
                        'border_color': '#D9D9D9',
                        'bg_color': '#F2F2F2',
                        'align': 'left',
                        'valign': 'vcenter',
                    })
                    for row_idx, row in enumerate(df.iter_rows()):
                        row_format = alt_cell_format if row_idx % 2 == 1 else cell_format
                        for col_idx, value in enumerate(row):
                            worksheet.write(row_idx + 1, col_idx, value, row_format)
                
                # 自动调整列宽（仅采样前 200 行）
                sample_rows = list(df.head(min(200, row_count)).iter_rows())
                for col_idx, col_name in enumerate(columns):
                    max_len = len(str(col_name))
                    for r in sample_rows:
                        v = r[col_idx]
                        l = len(str(v)) if v is not None else 0
                        if l > max_len:
                            max_len = l
                    worksheet.set_column(col_idx, col_idx, min(max_len + 4, 50))
                
                worksheet.set_row(0, 25)
                worksheet.freeze_panes(1, 0)
            finally:
                workbook.close()
        
    except ImportError as e:
        # 如果没有xlsxwriter或openpyxl，使用polars默认导出
        print(f"[警告] 缺少必要的库: {e}，使用polars默认导出")
        df.write_excel(filepath)


class DataExporter:
    """数据导出器"""
    
    def __init__(self, output_dir: str = "./data"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(self, data: list[dict], filename: Optional[str] = None) -> str:
        """导出数据到Excel（带样式）"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        if not data:
            # 创建空文件
            pl.DataFrame().write_excel(str(filepath))
        else:
            df = pl.DataFrame(data)
            _write_styled_excel(df, str(filepath))
        
        return str(filepath)
    
    def export_to_csv(self, data: list[dict], filename: Optional[str] = None) -> str:
        """导出数据到CSV"""
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"data_{timestamp}.csv"
        
        filepath = self.output_dir / filename
        
        if not data:
            # 创建空文件
            pl.DataFrame().write_csv(str(filepath))
        else:
            df = pl.DataFrame(data)
            df.write_csv(str(filepath))
        
        return str(filepath)
